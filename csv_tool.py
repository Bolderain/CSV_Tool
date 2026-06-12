# csv_tool.py
# PySide6 GUI version (replaces tkinter)
# Features:
# - Input: .csv (comma or semicolon) and .xlsx (multi-sheet support)
# - Output: always comma-separated CSV
# - Modes: Repeater / Headend / Proxie
# - Device type variants: R310/R320/R330, M300/M310/M320, P300/P310
# - Config name validation: warns if template name suggests different device type
# - Multi-sheet Excel: sheet selector in preview section
# - No-header support: checkbox for sheets without header rows
# - Presets stored in ./presets/presets.json (next to this script)
# - Customer mappings stored in ./presets/mappings.json
# - Auto preview on selection/add (first 5 rows), shows delimiter/sheet, headers, and detected mapping
# - Manual mapping dropdowns + "Save mapping for customer"
# - Skips empty rows and logs warnings (does not fail)
# - Self-bootstraps dependencies (PySide6, openpyxl) into a private venv in C:\ct\.venv

# -----------------------------
# Dependency bootstrap
# -----------------------------
import os
import sys
import subprocess
import shutil
from pathlib import Path

REQUIRED_PACKAGES = [
    "openpyxl",
    "PySide6-Essentials==6.10.2",
]


def _run(cmd: list[str]) -> None:
    subprocess.check_call(cmd)


def _real(p: Path) -> Path:
    return Path(os.path.realpath(str(p)))


def _venv_dir() -> Path:
    root = os.environ.get("SystemDrive", "C:")
    return _real(Path(root + r"\ct") / ".venv")


def _venv_python(venv_path: Path) -> Path:
    return venv_path / "Scripts" / "python.exe"


def _ensure_deps():
    try:
        import openpyxl  # noqa
        from PySide6 import QtWidgets  # noqa

        return
    except ModuleNotFoundError:
        pass

    venv_path = _real(_venv_dir())
    venv_path.parent.mkdir(parents=True, exist_ok=True)

    py = _venv_python(venv_path)
    cfg = venv_path / "pyvenv.cfg"

    if not py.exists() or not cfg.exists():
        if venv_path.exists():
            shutil.rmtree(venv_path, ignore_errors=True)
        _run([sys.executable, "-m", "venv", str(venv_path)])

        venv_path = _real(venv_path)
        py = _venv_python(venv_path)
        cfg = venv_path / "pyvenv.cfg"

        if not py.exists() or not cfg.exists():
            raise RuntimeError(f"Venv creation failed: missing pyvenv.cfg at {cfg}")

    _run([str(py), "-m", "pip", "install", "--upgrade", "pip"])
    _run([str(py), "-m", "pip", "install", *REQUIRED_PACKAGES])

    os.execv(str(py), [str(py)] + sys.argv)


_ensure_deps()

# -----------------------------
# Imports (after deps)
# -----------------------------
import csv
import json
import datetime as dt
import re

from openpyxl import load_workbook

from PySide6.QtCore import Qt, QSignalBlocker
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QGroupBox,
    QVBoxLayout,
    QHBoxLayout,
    QGridLayout,
    QListWidget,
    QListWidgetItem,
    QPushButton,
    QFileDialog,
    QMessageBox,
    QLabel,
    QLineEdit,
    QComboBox,
    QPlainTextEdit,
    QScrollArea,
    QSizePolicy,
    QCheckBox,
)


# -----------------------------
# Output schemas
# -----------------------------
REPEATER_OUTPUT_FIELDS = [
    "serialNumber",
    "macAddress",
    "type",
    "registrationStatus",
    "desiredConfigurationTemplate",
    "desiredConfigurationMd5",
    "desiredConfigurationSize",
]

HEADEND_OUTPUT_FIELDS = [
    "serialNumber",
    "macAddress",
    "type",
    "registrationStatus",
    "desiredConfigurationTemplate",
    "desiredConfigurationMd5",
    "desiredConfigurationSize",
]

PROXIE_OUTPUT_FIELDS = [
    "serialNumber",
    "macAddress",
    "type",
    "registrationStatus",
    "accessToken",
    "desiredConfigurationTemplate",
    "desiredConfigurationMd5",
    "desiredConfigurationSize",
]

# Device type variants per mode (editable combo — user can also type any custom value)
DEVICE_VARIANTS = {
    "Repeater": ["R310", "R320", "R330"],
    "Headend":  ["M200", "M300", "M310", "M320", "M400", "M500"],
    "Proxie":   ["P200", "P300", "P310", "P400", "P500"],
}

# Maps the leading letter of a device-type token (e.g. M400) to its mode
_DEVICE_LETTER_MODE = {"R": "Repeater", "M": "Headend", "P": "Proxie"}

ACCESS_TOKEN_PREFIX = "00185803"

SERIAL_HEADER_ALIASES = [
    "serialnumber",
    "serialno",
    "serialnr",
    "serialnum",
    "serial",
    "sn",
    "sno",
    "deviceserialnumber",
    "devserialnumber",
    "deviceid",
    "devicenumber",
    "seriennummer",
    "seriennr",
    "sernr",
    "idserial",
    "meterserial",
]

MAC_HEADER_ALIASES = [
    "macaddress",
    "macadress",
    "macaddr",
    "mac",
    "macid",
    "maceth",
    "hwaddress",
    "hardwareaddress",
    "ethernetaddress",
    "ethmac",
    "lanmac",
    "wlanmac",
    "eui",
    "eui64",
    "mac_eth",
    "macethernet",
]


# -----------------------------
# Paths: presets and mappings stored next to script
# -----------------------------
def _script_dir() -> Path:
    return Path(__file__).resolve().parent


def _preset_folder() -> Path:
    p = _script_dir() / "presets"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _presets_json_path() -> Path:
    return _preset_folder() / "presets.json"


def _mappings_json_path() -> Path:
    return _preset_folder() / "mappings.json"


# -----------------------------
# Helpers
# -----------------------------
def _norm_header(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").strip().lower())


def _pick_column(fieldnames: list[str], aliases: list[str], must_contain: list[str]) -> str | None:
    norm_map = {_norm_header(fn): fn for fn in fieldnames}
    for a in aliases:
        if a in norm_map:
            return norm_map[a]
    for fn in fieldnames:
        n = _norm_header(fn)
        if all(token in n for token in must_contain):
            return fn
    return None


def _sanitize_filename_part(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", s)
    s = s.strip().strip(".")
    return s or "input"


def _validate_yyyymmdd(s: str) -> bool:
    return bool(re.fullmatch(r"\d{8}", (s or "").strip()))


def _final_output_name(date_yyyymmdd: str, inputname: str, data_rows: int, device_type: str) -> str:
    date_part = _sanitize_filename_part(date_yyyymmdd)
    name_part = _sanitize_filename_part(inputname)
    type_part = _sanitize_filename_part(device_type)
    return f"Device_import_{date_part}_{name_part}_{data_rows}_{type_part}.csv"


def _validate_md5_hex32(s: str) -> bool:
    return bool(re.fullmatch(r"[0-9a-fA-F]{32}", (s or "").strip()))


def _validate_size_numeric(s: str) -> bool:
    return bool(re.fullmatch(r"\d+", (s or "").strip()))


def _detect_input_delimiter(file_obj) -> str:
    sample = file_obj.read(8192)
    file_obj.seek(0)

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        if getattr(dialect, "delimiter", None) in (",", ";"):
            return dialect.delimiter
    except csv.Error:
        pass

    lines = sample.splitlines()
    lines = [l for l in lines if l.strip()][:10]
    text = "\n".join(lines)
    return ";" if text.count(";") > text.count(",") else ","


def _is_empty_row(row: dict) -> bool:
    return all(((v or "").strip() == "") for v in row.values())


def _scan_all_rows(
    infile: Path,
    serial_col: str | None,
    sheet_name: str | None = None,
    no_header: bool = False,
) -> tuple[int, int, list[str]]:
    """Scan the whole file. Return (data_rows, empty_rows, dup_warning_msgs)."""
    total = 0
    empty = 0
    seen: dict[str, int] = {}
    dups: list[str] = []

    def _noop(_msg): pass

    for row_index, _fields, row in _iter_input_rows(infile, _noop, sheet_name=sheet_name, no_header=no_header):
        if _is_empty_row(row):
            empty += 1
            continue
        total += 1
        if serial_col:
            sn = (row.get(serial_col) or "").strip()
            if sn:
                if sn in seen:
                    if len(dups) < 20:
                        dups.append(f"Row {row_index}: '{sn}' (also row {seen[sn]})")
                else:
                    seen[sn] = row_index
    return total, empty, dups


def _validate_export(
    infile: Path,
    mode: str,
    run_cfg: dict,
    mapping_override: dict | None,
    sheet_name: str | None,
    no_header: bool,
    log_fn,
) -> dict:
    """Full row-by-row validation pass. Returns {ok_count, errors, skipped_empty, total_data}."""
    it = _iter_input_rows(infile, log_fn, sheet_name=sheet_name, no_header=no_header)
    first = next(it, None)
    if first is None:
        return {"ok_count": 0, "errors": [], "skipped_empty": 0, "total_data": 0}

    first_row_index, input_fields, first_row = first
    factory, _ = _FACTORY_BY_MODE[mode]

    try:
        transform = factory(input_fields, run_cfg, mapping_override)
    except ValueError as e:
        return {"ok_count": 0, "errors": [(0, str(e))], "skipped_empty": 0, "total_data": 0}

    ok_count = 0
    errors: list[tuple[int, str]] = []
    skipped_empty = 0
    seen_serials: dict[str, int] = {}

    def _process(row_index: int, row: dict):
        nonlocal ok_count, skipped_empty
        if _is_empty_row(row):
            skipped_empty += 1
            return
        try:
            result = transform(row, row_index)
            sn = result.get("serialNumber", "")
            if sn and sn in seen_serials:
                errors.append((row_index, f"Duplicate serialNumber '{sn}' (also at row {seen_serials[sn]})"))
            else:
                if sn:
                    seen_serials[sn] = row_index
                ok_count += 1
        except ValueError as e:
            errors.append((row_index, str(e)))

    _process(first_row_index, first_row)
    for row_index, _fields, row in it:
        _process(row_index, row)

    return {
        "ok_count": ok_count,
        "errors": errors,
        "skipped_empty": skipped_empty,
        "total_data": ok_count + len(errors),
    }


def _serial_c_to_b(serial: str) -> str:
    s = (serial or "").strip()
    if not s:
        raise ValueError("Headend mode: serialNumber is empty.")
    if s[0] in ("C", "c"):
        return "B" + s[1:]
    if s[0] in ("B", "b"):
        return "B" + s[1:]
    raise ValueError(f"Headend mode: serialNumber does not start with C or B: '{s}'")


def _mac_minus_one(mac: str) -> str:
    raw = re.sub(r"[^0-9a-fA-F]", "", (mac or "").strip())
    if len(raw) != 12:
        raise ValueError(f"Headend mode: invalid MAC (expected 12 hex digits): '{mac}'")
    value = int(raw, 16)
    if value == 0:
        raise ValueError(f"Headend mode: MAC underflow on decrement: '{mac}'")
    value -= 1
    h = f"{value:012X}"
    return ":".join(h[i : i + 2] for i in range(0, 12, 2))


def _normalize_mac_colonsep(mac: str) -> str:
    raw = re.sub(r"[^0-9a-fA-F]", "", (mac or "").strip())
    if len(raw) != 12:
        raise ValueError(f"Invalid MAC (expected 12 hex digits): '{mac}'")
    return ":".join(raw[i : i + 2].upper() for i in range(0, 12, 2))


def _normalize_mac_12hex_lower(mac: str) -> str:
    raw = re.sub(r"[^0-9a-fA-F]", "", (mac or "").strip())
    if len(raw) != 12:
        raise ValueError(f"Proxie mode: invalid MAC (expected 12 hex digits): '{mac}'")
    return raw.lower()


def _access_token_from_mac(mac: str) -> str:
    return ACCESS_TOKEN_PREFIX + _normalize_mac_12hex_lower(mac)


def _excel_col_letter(idx: int) -> str:
    """Convert 0-based column index to Excel column letter(s): 0→A, 25→Z, 26→AA …"""
    result = ""
    n = idx + 1
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(ord("A") + r) + result
    return result


def _get_xlsx_sheet_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _check_config_compatibility(name: str, mode: str, device_type: str, label: str = "Config") -> str | None:
    """Return a warning+suggestion string if name suggests a different device type/mode, else None.

    Detects any token [R|M|P]NNN (letter + exactly 3 digits) — future-proof for M400, P500 etc.
    """
    if not name:
        return None
    t = name.upper()

    # Match any [R/M/P] followed by exactly 3 digits, not part of a longer alphanumeric token
    m = re.search(r"(?<![A-Z0-9])([RMP])(\d{3})(?![A-Z0-9])", t)
    if m:
        found_dtype = m.group(0)  # e.g. "M400"
        expected_mode = _DEVICE_LETTER_MODE[m.group(1)]
        if expected_mode != mode:
            return (
                f"{label} contains '{found_dtype}' ({expected_mode}) but mode is {mode}. "
                f"→ Switch to {expected_mode} mode."
            )
        if found_dtype != device_type.upper():
            return (
                f"{label} contains '{found_dtype}' but device type is set to {device_type}. "
                f"→ Change device type to {found_dtype}."
            )
        return None  # exact match, all good

    # Mode-family keyword fallback (for names without a numeric device token)
    if re.search(r"(?<![A-Z])HE(?![A-Z])|HEADEND", t) and mode != "Headend":
        return (
            f"{label} suggests Headend (HE/HEADEND) but mode is {mode}. "
            f"→ Switch to Headend mode."
        )
    if re.search(r"PROXY|PROX(?![A-Z])", t) and mode != "Proxie":
        return (
            f"{label} suggests Proxie (PROXY) but mode is {mode}. "
            f"→ Switch to Proxie mode."
        )
    if re.search(r"EBRO", t) and mode != "Repeater":
        return (
            f"{label} suggests Repeater (EBRO) but mode is {mode}. "
            f"→ Switch to Repeater mode."
        )

    return None


def _build_run_cfg(mode: str, device_type: str, template: str, md5: str, size: str) -> dict:
    if mode not in DEVICE_VARIANTS:
        raise ValueError(f"Unknown mode: {mode}")

    device_type = (device_type or "").strip()
    if not device_type:
        raise ValueError(
            f"Device type is empty. Pick a preset ({', '.join(DEVICE_VARIANTS[mode])}) "
            "or type your own."
        )

    template = (template or "").strip()
    md5 = (md5 or "").strip()
    size = (size or "").strip()

    if not template:
        raise ValueError("desiredConfigurationTemplate is empty.")
    if not md5:
        raise ValueError("desiredConfigurationMd5 is empty.")
    if not size:
        raise ValueError("desiredConfigurationSize is empty.")
    if not _validate_md5_hex32(md5):
        raise ValueError("desiredConfigurationMd5 must be 32 hex characters.")
    if not _validate_size_numeric(size):
        raise ValueError("desiredConfigurationSize must be numeric.")

    return {
        "type": device_type,
        "registrationStatus": "ACTIVATED",
        "desiredConfigurationTemplate": template,
        "desiredConfigurationMd5": md5.lower(),
        "desiredConfigurationSize": size,
    }


# -----------------------------
# XLSX reading helpers
# -----------------------------
def _build_xlsx_fieldnames(raw_header: tuple, no_header: bool) -> list[str]:
    """
    Build column name list from a raw header row.
    If no_header=True every column gets a positional name.
    If no_header=False empty cells get a positional fallback (Col_A, Col_B …).
    """
    if no_header:
        return [f"Col_{_excel_col_letter(i)}" for i in range(len(raw_header))]
    result = []
    for i, h in enumerate(raw_header):
        name = str(h).strip() if h is not None else ""
        if not name:
            name = f"Col_{_excel_col_letter(i)}"
        result.append(name)
    return result


def _read_xlsx_headers_and_rows(
    path: Path,
    sheet_name: str | None = None,
    no_header: bool = False,
    max_rows: int | None = None,
):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        it = ws.iter_rows(values_only=True)
        first = next(it, None)
        if first is None:
            raise ValueError(f"No data in sheet '{ws.title}' in file: {path}")

        fieldnames = _build_xlsx_fieldnames(first, no_header)
        num_cols = len(fieldnames)

        rows = []
        start_row = 1 if no_header else 2

        if no_header:
            # first row is a data row
            row0 = {
                fieldnames[i]: ("" if i >= len(first) or first[i] is None else str(first[i]).strip())
                for i in range(num_cols)
            }
            rows.append((1, row0))

        for excel_row_num, values in enumerate(it, start=start_row + (0 if no_header else 0)):
            row = {
                fieldnames[i]: ("" if i >= len(values) or values[i] is None else str(values[i]).strip())
                for i in range(num_cols)
            }
            rows.append((excel_row_num if no_header else excel_row_num, row))
            if max_rows is not None and len(rows) >= max_rows:
                break

        return ws.title, fieldnames, rows
    finally:
        wb.close()


def _read_input_preview(
    infile: Path,
    sheet_name: str | None = None,
    no_header: bool = False,
    max_rows: int = 5,
) -> dict:
    suffix = infile.suffix.lower()

    if suffix == ".csv":
        with infile.open("r", encoding="utf-8-sig", newline="") as fin:
            delim = _detect_input_delimiter(fin)
            reader = csv.DictReader(fin, delimiter=delim)
            if reader.fieldnames is None:
                raise ValueError(f"No header found in file: {infile}")
            headers = list(reader.fieldnames)
            rows = []
            for row_index, row in enumerate(reader, start=2):
                rows.append((row_index, row))
                if len(rows) >= max_rows:
                    break
            return {
                "type": "csv",
                "delimiter": delim,
                "sheet": "",
                "headers": headers,
                "rows": rows,
                "all_sheets": [],
            }

    if suffix == ".xlsx":
        all_sheets = _get_xlsx_sheet_names(infile)
        sheet, headers, rows = _read_xlsx_headers_and_rows(
            infile, sheet_name=sheet_name, no_header=no_header, max_rows=max_rows
        )
        return {
            "type": "xlsx",
            "delimiter": "",
            "sheet": sheet,
            "headers": headers,
            "rows": rows,
            "all_sheets": all_sheets,
        }

    raise ValueError(f"Unsupported input type: {infile.suffix}")


def _iter_input_rows(
    infile: Path,
    log_fn,
    sheet_name: str | None = None,
    no_header: bool = False,
):
    suffix = infile.suffix.lower()

    if suffix == ".csv":
        with infile.open("r", encoding="utf-8-sig", newline="") as fin:
            in_delim = _detect_input_delimiter(fin)
            log_fn(f"Detected input delimiter: '{in_delim}' for {infile.name}")
            reader = csv.DictReader(fin, delimiter=in_delim)
            if reader.fieldnames is None:
                raise ValueError(f"No header found in file: {infile}")
            fieldnames = list(reader.fieldnames)
            for row_index, row in enumerate(reader, start=2):
                yield row_index, fieldnames, row
        return

    if suffix == ".xlsx":
        wb = load_workbook(infile, read_only=True, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            log_fn(f"Detected input type: xlsx (sheet '{ws.title}') for {infile.name}")
            it = ws.iter_rows(values_only=True)
            first = next(it, None)
            if first is None:
                raise ValueError(f"No data in sheet '{ws.title}' in file: {infile}")

            fieldnames = _build_xlsx_fieldnames(first, no_header)
            num_cols = len(fieldnames)

            if not no_header and all(fn.startswith("Col_") for fn in fieldnames):
                log_fn(
                    f"WARNING: All column headers are empty in sheet '{ws.title}'. "
                    "Consider enabling 'No header row' if the sheet has no header."
                )

            if no_header:
                # first row IS a data row
                row0 = {
                    fieldnames[i]: ("" if i >= len(first) or first[i] is None else str(first[i]).strip())
                    for i in range(num_cols)
                }
                yield 1, fieldnames, row0

            for excel_row_num, values in enumerate(it, start=2):
                row = {
                    fieldnames[i]: ("" if i >= len(values) or values[i] is None else str(values[i]).strip())
                    for i in range(num_cols)
                }
                yield excel_row_num, fieldnames, row
        finally:
            wb.close()
        return

    raise ValueError(f"Unsupported input type: {infile.suffix}")


# -----------------------------
# Presets / Mappings storage
# -----------------------------
def _default_presets_payload() -> dict:
    return {
        "Repeater": {
            "DEFAULT": {
                "desiredConfigurationTemplate": "CFG-0001-0100-EBRO-20230223-FTPS-v1.tar.gz",
                "desiredConfigurationMd5": "f09ffd00e02bf9ad2108ea5199d46d2c",
                "desiredConfigurationSize": "195",
            }
        },
        "Headend": {
            "DEFAULT": {
                "desiredConfigurationTemplate": "CFG-0001-0100-HE-20230223-FTPS_AGC-v1.tar.gz",
                "desiredConfigurationMd5": "648be952a640f0d00f5da9f12854477f",
                "desiredConfigurationSize": "215",
            }
        },
        "Proxie": {
            "DEFAULT": {
                "desiredConfigurationTemplate": "CFG-0001-0100-PROXY-20230131-FTPS-v1.tar.gz",
                "desiredConfigurationMd5": "9768cbe43fa48cec894443e8c03ed399",
                "desiredConfigurationSize": "630",
            }
        },
    }


def _load_presets() -> dict:
    path = _presets_json_path()
    if not path.exists():
        payload = _default_presets_payload()
        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        return payload

    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    out = {"Repeater": {}, "Headend": {}, "Proxie": {}}
    for mode in out.keys():
        v = data.get(mode, {})
        if isinstance(v, dict):
            out[mode] = v
    return out


def _save_presets(presets: dict) -> None:
    path = _presets_json_path()
    payload = {
        "Repeater": presets.get("Repeater", {}),
        "Headend": presets.get("Headend", {}),
        "Proxie": presets.get("Proxie", {}),
    }
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def _load_mappings() -> dict:
    path = _mappings_json_path()
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _save_mappings(mappings: dict) -> None:
    path = _mappings_json_path()
    with path.open("w", encoding="utf-8") as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)


# -----------------------------
# Transform factories with mapping override
# -----------------------------
def _resolve_serial_mac_keys(
    input_fieldnames: list[str], mapping_override: dict | None
) -> tuple[str | None, str | None, str]:
    if mapping_override:
        s = (mapping_override.get("serial") or "").strip()
        m = (mapping_override.get("mac") or "").strip()
        if s and m and s in input_fieldnames and m in input_fieldnames:
            return s, m, "manual"

    serial_key = _pick_column(input_fieldnames, SERIAL_HEADER_ALIASES, must_contain=["serial"])
    mac_key = _pick_column(input_fieldnames, MAC_HEADER_ALIASES, must_contain=["mac"])

    return serial_key, mac_key, "auto"


def _repeater_transform_factory(
    input_fieldnames: list[str], run_cfg: dict, mapping_override: dict | None
):
    serial_key, mac_key, _src = _resolve_serial_mac_keys(input_fieldnames, mapping_override)
    if not serial_key or not mac_key:
        raise ValueError(
            "Repeater mode: could not find required columns (serial and mac). "
            "Use Mapping dropdowns and save mapping for the customer."
        )

    def transform(row: dict, row_index: int) -> dict:
        serial = (row.get(serial_key) or "").strip()
        mac = (row.get(mac_key) or "").strip()
        if not serial:
            raise ValueError(
                f"Repeater mode: empty serialNumber at row {row_index} (column: '{serial_key}')"
            )
        if not mac:
            raise ValueError(
                f"Repeater mode: empty macAddress at row {row_index} (column: '{mac_key}')"
            )
        return {
            "serialNumber": serial,
            "macAddress": mac,
            "type": run_cfg["type"],
            "registrationStatus": run_cfg["registrationStatus"],
            "desiredConfigurationTemplate": run_cfg["desiredConfigurationTemplate"],
            "desiredConfigurationMd5": run_cfg["desiredConfigurationMd5"],
            "desiredConfigurationSize": run_cfg["desiredConfigurationSize"],
        }

    return transform


def _headend_transform_factory(
    input_fieldnames: list[str], run_cfg: dict, mapping_override: dict | None
):
    serial_key, mac_key, _src = _resolve_serial_mac_keys(input_fieldnames, mapping_override)
    if not serial_key or not mac_key:
        raise ValueError(
            "Headend mode: could not find required columns (serial and mac). "
            "Use Mapping dropdowns and save mapping for the customer."
        )

    def transform(row: dict, row_index: int) -> dict:
        in_serial = (row.get(serial_key) or "").strip()
        in_mac = (row.get(mac_key) or "").strip()
        if not in_serial:
            raise ValueError(
                f"Headend mode: empty serialNumber at row {row_index} (column: '{serial_key}')"
            )
        if not in_mac:
            raise ValueError(
                f"Headend mode: empty macAddress at row {row_index} (column: '{mac_key}')"
            )
        out_serial = _serial_c_to_b(in_serial)
        out_mac = _mac_minus_one(in_mac)
        return {
            "serialNumber": out_serial,
            "macAddress": out_mac,
            "type": run_cfg["type"],
            "registrationStatus": run_cfg["registrationStatus"],
            "desiredConfigurationTemplate": run_cfg["desiredConfigurationTemplate"],
            "desiredConfigurationMd5": run_cfg["desiredConfigurationMd5"],
            "desiredConfigurationSize": run_cfg["desiredConfigurationSize"],
        }

    return transform


def _proxie_transform_factory(
    input_fieldnames: list[str], run_cfg: dict, mapping_override: dict | None
):
    serial_key, mac_key, _src = _resolve_serial_mac_keys(input_fieldnames, mapping_override)
    if not serial_key or not mac_key:
        raise ValueError(
            "Proxie mode: could not find required columns (serial and mac). "
            "Use Mapping dropdowns and save mapping for the customer."
        )

    def transform(row: dict, row_index: int) -> dict:
        serial = (row.get(serial_key) or "").strip()
        mac_in = (row.get(mac_key) or "").strip()
        if not serial:
            raise ValueError(
                f"Proxie mode: empty serialNumber at row {row_index} (column: '{serial_key}')"
            )
        if not mac_in:
            raise ValueError(
                f"Proxie mode: empty macAddress at row {row_index} (column: '{mac_key}')"
            )
        normalized_mac = _normalize_mac_colonsep(mac_in)
        token = _access_token_from_mac(mac_in)
        return {
            "serialNumber": serial,
            "macAddress": normalized_mac,
            "type": run_cfg["type"],
            "registrationStatus": run_cfg["registrationStatus"],
            "accessToken": token,
            "desiredConfigurationTemplate": run_cfg["desiredConfigurationTemplate"],
            "desiredConfigurationMd5": run_cfg["desiredConfigurationMd5"],
            "desiredConfigurationSize": run_cfg["desiredConfigurationSize"],
        }

    return transform


# -----------------------------
# GUI
# -----------------------------
class CsvToolWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CSV Tool")
        self.resize(1160, 960)
        self.setMinimumSize(1000, 820)

        self.script_dir = _script_dir()
        self.output_dir = self.script_dir

        self.files: list[Path] = []

        self.presets = _load_presets()
        self.mappings = _load_mappings()

        self.mode_edit = "Repeater"

        self._building = False
        self._applying_preset = False
        self._per_mode_selected_preset = {
            "Repeater": self._default_preset_name("Repeater"),
            "Headend": self._default_preset_name("Headend"),
            "Proxie": self._default_preset_name("Proxie"),
        }
        self._per_mode_device_type = {
            mode: variants[0] for mode, variants in DEVICE_VARIANTS.items()
        }
        self._xlsx_sheets: list[str] = []

        self._build_ui()
        self._log(f"Presets file: {_presets_json_path()}")
        self._log(f"Mappings file: {_mappings_json_path()}")

        self._set_mode_edit("Repeater")

    def _default_preset_name(self, mode: str) -> str:
        block = self.presets.get(mode, {})
        if isinstance(block, dict) and block:
            if "DEFAULT" in block:
                return "DEFAULT"
            return sorted(block.keys())[0]
        return "DEFAULT"

    def _build_ui(self):
        self._building = True

        central = QWidget()
        root_layout = QVBoxLayout(central)
        root_layout.setSpacing(10)
        root_layout.setContentsMargins(10, 10, 10, 10)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        self.form_layout = QVBoxLayout(container)
        self.form_layout.setSpacing(10)
        self.form_layout.setContentsMargins(10, 10, 10, 10)
        scroll.setWidget(container)

        root_layout.addWidget(scroll)
        self.setCentralWidget(central)

        # -------------------------
        # Input files group
        # -------------------------
        gb_files = QGroupBox("Input files (.csv or .xlsx)")
        files_layout = QGridLayout(gb_files)
        files_layout.setColumnStretch(0, 1)

        self.file_list = QListWidget()
        self.file_list.setToolTip("Input files list. Selecting a file auto-previews it.")
        self.file_list.setFixedHeight(90)
        self.file_list.currentRowChanged.connect(self._on_file_selected)

        files_layout.addWidget(self.file_list, 0, 0, 4, 1)

        btn_add = QPushButton("Add files")
        btn_add.setToolTip(
            "Add one or more input files (.csv or .xlsx). Newly added file is auto-selected and previewed."
        )
        btn_add.clicked.connect(self._add_files)

        btn_remove = QPushButton("Remove selected")
        btn_remove.setToolTip("Remove the selected file from the list.")
        btn_remove.clicked.connect(self._remove_selected)

        btn_clear = QPushButton("Clear list")
        btn_clear.setToolTip("Remove all files from the list.")
        btn_clear.clicked.connect(self._clear_list)

        btn_preview = QPushButton("Preview selected")
        btn_preview.setToolTip(
            "Show the first 5 rows, detected delimiter/sheet, headers, and mapping."
        )
        btn_preview.clicked.connect(self._preview_selected)

        files_layout.addWidget(btn_add, 0, 1)
        files_layout.addWidget(btn_remove, 1, 1)
        files_layout.addWidget(btn_clear, 2, 1)
        files_layout.addWidget(btn_preview, 3, 1)

        self.form_layout.addWidget(gb_files)

        # -------------------------
        # Output and filename group
        # -------------------------
        gb_out = QGroupBox("Output")
        out_layout = QGridLayout(gb_out)
        out_layout.setColumnStretch(1, 1)

        out_layout.addWidget(QLabel("Output folder"), 0, 0)
        self.out_dir = QLineEdit(str(self.output_dir))
        self.out_dir.setReadOnly(True)
        out_layout.addWidget(self.out_dir, 0, 1)

        btn_choose = QPushButton("Choose")
        btn_choose.setToolTip("Select the output folder. Output is always comma-separated CSV.")
        btn_choose.clicked.connect(self._choose_output_dir)
        out_layout.addWidget(btn_choose, 0, 2)

        out_layout.addWidget(QLabel("Filename date (YYYYMMDD)"), 1, 0)
        self.date_edit = QLineEdit(dt.datetime.now().strftime("%Y%m%d"))
        self.date_edit.setToolTip("Date used in output filename.")
        self.date_edit.setMaximumWidth(120)
        out_layout.addWidget(self.date_edit, 1, 1, alignment=Qt.AlignLeft)

        out_layout.addWidget(QLabel("INPUTNAME_CUSTOMER"), 1, 1, alignment=Qt.AlignCenter)
        self.customer_edit = QLineEdit("")
        self.customer_edit.setToolTip(
            "Customer name used in output filename and customer mapping key."
        )
        self.customer_edit.setMaximumWidth(260)
        out_layout.addWidget(self.customer_edit, 1, 2, alignment=Qt.AlignLeft)

        hint = QLabel("Output filename: Device_import_YYYYMMDD_INPUTNAME_CUSTOMER_ROWCOUNT_TYPE.csv")
        hint.setStyleSheet("color: gray;")
        out_layout.addWidget(hint, 2, 0, 1, 3)

        self.form_layout.addWidget(gb_out)

        # -------------------------
        # Desired configuration group
        # -------------------------
        gb_cfg = QGroupBox("Desired configuration (presets + editable)")
        cfg_layout = QGridLayout(gb_cfg)
        cfg_layout.setColumnStretch(7, 1)

        # Row 0: Edit for | combo | Device type | combo | Preset | combo | New preset name | field | Save
        cfg_layout.addWidget(QLabel("Edit for"), 0, 0)
        self.edit_for = QComboBox()
        self.edit_for.addItems(["Repeater", "Headend", "Proxie"])
        self.edit_for.setToolTip(
            "Select which device type you want to edit desired configuration for."
        )
        self.edit_for.currentTextChanged.connect(self._set_mode_edit)
        cfg_layout.addWidget(self.edit_for, 0, 1)

        cfg_layout.addWidget(QLabel("Device type"), 0, 2)
        self.device_type_combo = QComboBox()
        self.device_type_combo.setEditable(True)
        self.device_type_combo.setInsertPolicy(QComboBox.NoInsert)
        self.device_type_combo.setMinimumWidth(110)
        self.device_type_combo.setToolTip(
            "Select a device type variant (e.g. R310, R320, R330 for Repeater),\n"
            "or type your own value. This determines the 'type' field in the output CSV."
        )
        self.device_type_combo.currentTextChanged.connect(self._on_device_type_changed)
        cfg_layout.addWidget(self.device_type_combo, 0, 3)

        cfg_layout.addWidget(QLabel("Preset"), 0, 4)
        self.preset_combo = QComboBox()
        self.preset_combo.setToolTip(
            "Select a preset to fill Template/MD5/Size. Editing fields switches to Custom."
        )
        self.preset_combo.currentTextChanged.connect(self._on_preset_changed)
        cfg_layout.addWidget(self.preset_combo, 0, 5)

        cfg_layout.addWidget(QLabel("New preset name"), 0, 6)
        self.new_preset_name = QLineEdit("")
        self.new_preset_name.setToolTip(
            "Name for a new preset. Must be unique within the selected mode."
        )
        cfg_layout.addWidget(self.new_preset_name, 0, 7)

        btn_save_preset = QPushButton("Save preset")
        btn_save_preset.setToolTip(
            "Save the current Template/MD5/Size as a new preset for this mode."
        )
        btn_save_preset.clicked.connect(self._save_preset)
        cfg_layout.addWidget(btn_save_preset, 0, 8)

        # Row 1: template
        cfg_layout.addWidget(QLabel("desiredConfigurationTemplate"), 1, 0)
        self.template_edit = QLineEdit("")
        self.template_edit.textChanged.connect(self._mark_custom_if_user_edit)
        cfg_layout.addWidget(self.template_edit, 1, 1, 1, 8)

        # Row 2: config compatibility warning (empty = hidden visually via style)
        self.config_warn_label = QLabel("")
        self.config_warn_label.setStyleSheet(
            "color: #cc6600; font-weight: bold; padding: 0px 0px 2px 0px;"
        )
        self.config_warn_label.setVisible(False)
        cfg_layout.addWidget(self.config_warn_label, 2, 0, 1, 9)

        # Row 3: md5
        cfg_layout.addWidget(QLabel("desiredConfigurationMd5"), 3, 0)
        self.md5_edit = QLineEdit("")
        self.md5_edit.textChanged.connect(self._mark_custom_if_user_edit)
        cfg_layout.addWidget(self.md5_edit, 3, 1, 1, 8)

        # Row 4: size
        cfg_layout.addWidget(QLabel("desiredConfigurationSize"), 4, 0)
        self.size_edit = QLineEdit("")
        self.size_edit.setMaximumWidth(120)
        self.size_edit.textChanged.connect(self._mark_custom_if_user_edit)
        cfg_layout.addWidget(self.size_edit, 4, 1, alignment=Qt.AlignLeft)

        self.form_layout.addWidget(gb_cfg)

        # -------------------------
        # Preview + mapping group
        # -------------------------
        gb_preview = QGroupBox("Input preview (first 5 rows) and mapping")
        pv_layout = QVBoxLayout(gb_preview)

        self.preview_info = QLabel("No file previewed.")
        self.preview_info.setStyleSheet("color: gray;")
        self.mapping_info = QLabel("Mapping: none")
        self.mapping_info.setStyleSheet("color: gray;")

        pv_layout.addWidget(self.preview_info)
        pv_layout.addWidget(self.mapping_info)

        # Sheet selector row (only shown for multi-sheet xlsx)
        sheet_row = QHBoxLayout()
        self.sheet_label = QLabel("Sheet:")
        sheet_row.addWidget(self.sheet_label)
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(300)
        self.sheet_combo.setToolTip(
            "Select which sheet to preview and process. Only shown for multi-sheet Excel files."
        )
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        sheet_row.addWidget(self.sheet_combo)

        sheet_row.addSpacing(16)
        self.no_header_cb = QCheckBox("No header row (first row is data)")
        self.no_header_cb.setToolTip(
            "Check this when the sheet has no header row.\n"
            "Columns will be named Col_A, Col_B, etc.\n"
            "Set Serial/MAC mapping manually and save for this customer."
        )
        self.no_header_cb.toggled.connect(self._on_no_header_changed)
        sheet_row.addWidget(self.no_header_cb)
        sheet_row.addStretch(1)

        self.sheet_label.setVisible(False)
        self.sheet_combo.setVisible(False)

        pv_layout.addLayout(sheet_row)

        # Mapping row
        map_row = QHBoxLayout()
        map_row.addWidget(QLabel("Serial column"))
        self.serial_combo = QComboBox()
        self.serial_combo.setToolTip("Select which input column contains the serial number.")
        self.serial_combo.setMinimumWidth(240)
        map_row.addWidget(self.serial_combo)

        map_row.addSpacing(10)
        map_row.addWidget(QLabel("MAC column"))
        self.mac_combo = QComboBox()
        self.mac_combo.setToolTip("Select which input column contains the MAC address.")
        self.mac_combo.setMinimumWidth(240)
        map_row.addWidget(self.mac_combo)

        map_row.addSpacing(10)
        btn_save_map = QPushButton("Save mapping for customer")
        btn_save_map.setToolTip(
            "Save selected Serial/MAC columns for the current customer (Filename input name)."
        )
        btn_save_map.clicked.connect(self._save_mapping_for_customer)
        map_row.addWidget(btn_save_map)
        map_row.addStretch(1)

        pv_layout.addLayout(map_row)

        self.preview_text = QPlainTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setLineWrapMode(QPlainTextEdit.NoWrap)
        self.preview_text.setFixedHeight(140)
        pv_layout.addWidget(self.preview_text)

        self.form_layout.addWidget(gb_preview)

        # -------------------------
        # Scripts group
        # -------------------------
        gb_scripts = QGroupBox("Scripts")
        sc_layout = QHBoxLayout(gb_scripts)

        self.btn_rep = QPushButton("Repeater (R310)")
        self.btn_rep.setToolTip(
            "Generate output CSV for Repeaters. Uses the device type shown in the button label."
        )
        self.btn_rep.clicked.connect(lambda: self._run_mode("Repeater"))

        self.btn_he = QPushButton("Headend (M300)")
        self.btn_he.setToolTip(
            "Generate output CSV for Headends. Serial C→B, MAC−1. "
            "Uses the device type shown in the button label."
        )
        self.btn_he.clicked.connect(lambda: self._run_mode("Headend"))

        self.btn_px = QPushButton("Proxie (P300)")
        self.btn_px.setToolTip(
            "Generate output CSV for Proxies. Adds accessToken = 00185803 + MAC (no separators, lowercase). "
            "MAC output formatted as XX:XX:XX:XX:XX:XX. Uses the device type shown in the button label."
        )
        self.btn_px.clicked.connect(lambda: self._run_mode("Proxie"))

        self.status = QLabel("Ready.")
        self.status.setStyleSheet("color: gray;")

        sc_layout.addWidget(self.btn_rep)
        sc_layout.addWidget(self.btn_he)
        sc_layout.addWidget(self.btn_px)
        sc_layout.addSpacing(16)
        sc_layout.addWidget(self.status)
        sc_layout.addStretch(1)

        self.form_layout.addWidget(gb_scripts)

        # -------------------------
        # Log group
        # -------------------------
        gb_log = QGroupBox("Log")
        log_layout = QVBoxLayout(gb_log)

        self.log_text = QPlainTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setLineWrapMode(QPlainTextEdit.WidgetWidth)
        self.log_text.setFixedHeight(160)
        log_layout.addWidget(self.log_text)

        self.form_layout.addWidget(gb_log)

        self._building = False

    # -----------------------------
    # Logging
    # -----------------------------
    def _log(self, msg: str) -> None:
        ts = dt.datetime.now().strftime("%H:%M:%S")
        self.log_text.appendPlainText(f"[{ts}] {msg}")

    # -----------------------------
    # File list actions
    # -----------------------------
    def _add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Select input files",
            "",
            "CSV or Excel (*.csv *.xlsx);;CSV files (*.csv);;Excel files (*.xlsx);;All files (*.*)",
        )
        if not paths:
            return

        first_new_index = None
        added = 0

        for p in paths:
            fp = Path(p)
            if fp not in self.files:
                self.files.append(fp)
                added += 1
                if first_new_index is None:
                    first_new_index = len(self.files) - 1

        self._refresh_file_list()
        self._log(f"Added {added} file(s). Total: {len(self.files)}.")

        if self.files and first_new_index is not None:
            self.file_list.setCurrentRow(first_new_index)
            self._preview_selected()

    def _remove_selected(self):
        idx = self.file_list.currentRow()
        if idx < 0 or idx >= len(self.files):
            return
        removed = self.files.pop(idx)
        self._refresh_file_list()
        self._log(f"Removed: {removed.name}. Total: {len(self.files)}.")

        if self.files:
            self.file_list.setCurrentRow(min(idx, len(self.files) - 1))
            self._preview_selected()
        else:
            self._clear_preview()

    def _clear_list(self):
        self.files.clear()
        self._refresh_file_list()
        self._log("Cleared file list.")
        self._clear_preview()

    def _refresh_file_list(self):
        with QSignalBlocker(self.file_list):
            self.file_list.clear()
            for f in self.files:
                self.file_list.addItem(QListWidgetItem(str(f)))

    def _on_file_selected(self, _row: int):
        if not self._building:
            self._preview_selected()

    def _selected_file(self) -> Path | None:
        idx = self.file_list.currentRow()
        if idx < 0 or idx >= len(self.files):
            return None
        return self.files[idx]

    def _clear_preview(self):
        self.preview_info.setText("No file previewed.")
        self.mapping_info.setText("Mapping: none")
        self.preview_text.setPlainText("")
        with QSignalBlocker(self.serial_combo):
            self.serial_combo.clear()
        with QSignalBlocker(self.mac_combo):
            self.mac_combo.clear()
        self._xlsx_sheets = []
        self._set_sheet_selector_visible(False)

    def _set_sheet_selector_visible(self, visible: bool):
        self.sheet_label.setVisible(visible)
        self.sheet_combo.setVisible(visible)

    # -----------------------------
    # Output folder
    # -----------------------------
    def _choose_output_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Select output folder", str(self.output_dir))
        if not d:
            return
        self.output_dir = Path(d)
        self.out_dir.setText(str(self.output_dir))
        self._log(f"Output folder set: {self.output_dir}")

    # -----------------------------
    # Device type and config validation
    # -----------------------------
    def _on_device_type_changed(self, dtype: str):
        if self._building:
            return
        dtype = (dtype or "").strip()
        if not dtype:
            return
        self._per_mode_device_type[self.mode_edit] = dtype
        self._refresh_run_buttons()
        self._update_config_warning()

    def _refresh_run_buttons(self):
        self.btn_rep.setText(f"Repeater ({self._per_mode_device_type['Repeater']})")
        self.btn_he.setText(f"Headend ({self._per_mode_device_type['Headend']})")
        self.btn_px.setText(f"Proxie ({self._per_mode_device_type['Proxie']})")

    def _update_config_warning(self):
        if self._building:
            return
        template = (self.template_edit.text() or "").strip()
        mode = self.mode_edit
        device_type = self._per_mode_device_type.get(mode, "")

        warnings = []
        w = _check_config_compatibility(template, mode, device_type, label="Template")
        if w:
            warnings.append(w)
        f = self._selected_file()
        if f:
            w2 = _check_config_compatibility(f.stem, mode, device_type, label="File name")
            if w2:
                warnings.append(w2)

        if warnings:
            self.config_warn_label.setText("⚠  " + "  |  ".join(warnings))
            self.config_warn_label.setVisible(True)
        else:
            self.config_warn_label.setText("")
            self.config_warn_label.setVisible(False)

    # -----------------------------
    # Desired config preset logic
    # -----------------------------
    def _set_mode_edit(self, mode: str):
        self.mode_edit = mode
        self.presets = _load_presets()
        names = sorted(self.presets.get(mode, {}).keys())
        if "Custom" not in names:
            names.append("Custom")

        with QSignalBlocker(self.preset_combo):
            self.preset_combo.clear()
            self.preset_combo.addItems(names)
            selected = self._per_mode_selected_preset.get(mode, "DEFAULT")
            if selected not in names:
                selected = "DEFAULT" if "DEFAULT" in names else (names[0] if names else "Custom")
            self.preset_combo.setCurrentText(selected)

        # Update device type combo for this mode (editable: keeps custom values too)
        variants = DEVICE_VARIANTS.get(mode, [])
        with QSignalBlocker(self.device_type_combo):
            self.device_type_combo.clear()
            self.device_type_combo.addItems(variants)
            current_dtype = self._per_mode_device_type.get(mode, "") or (variants[0] if variants else "")
            self.device_type_combo.setCurrentText(current_dtype)
            self._per_mode_device_type[mode] = current_dtype

        self._apply_preset(mode, self.preset_combo.currentText())
        self._refresh_run_buttons()
        self._update_config_warning()

    def _apply_preset(self, mode: str, preset_name: str):
        if preset_name == "Custom":
            return
        block = self.presets.get(mode, {})
        values = block.get(preset_name, {})
        if not isinstance(values, dict):
            return

        self._applying_preset = True
        try:
            with QSignalBlocker(self.template_edit):
                self.template_edit.setText(
                    str(values.get("desiredConfigurationTemplate", "")).strip()
                )
            with QSignalBlocker(self.md5_edit):
                self.md5_edit.setText(str(values.get("desiredConfigurationMd5", "")).strip())
            with QSignalBlocker(self.size_edit):
                self.size_edit.setText(str(values.get("desiredConfigurationSize", "")).strip())
        finally:
            self._applying_preset = False

        self._update_config_warning()

    def _on_preset_changed(self, preset: str):
        if self._building:
            return
        self._per_mode_selected_preset[self.mode_edit] = preset
        self._apply_preset(self.mode_edit, preset)

    def _mark_custom_if_user_edit(self, _text: str):
        if self._building or self._applying_preset:
            return
        if self.preset_combo.currentText() != "Custom":
            with QSignalBlocker(self.preset_combo):
                if self.preset_combo.findText("Custom") >= 0:
                    self.preset_combo.setCurrentText("Custom")
            self._per_mode_selected_preset[self.mode_edit] = "Custom"
        self._update_config_warning()

    def _save_preset(self):
        mode = self.mode_edit
        name = (self.new_preset_name.text() or "").strip()

        if not name:
            QMessageBox.critical(self, "Preset name", "Preset name is empty.")
            return
        if name in ("DEFAULT", "Custom"):
            QMessageBox.critical(self, "Preset name", "Preset name cannot be DEFAULT or Custom.")
            return

        self.presets = _load_presets()
        if name in self.presets.get(mode, {}):
            QMessageBox.critical(
                self,
                "Preset name",
                f"Preset name already exists for {mode}. Choose another name.",
            )
            return

        template = (self.template_edit.text() or "").strip()
        md5 = (self.md5_edit.text() or "").strip()
        size = (self.size_edit.text() or "").strip()

        if not template:
            QMessageBox.critical(self, "Preset values", "desiredConfigurationTemplate is empty.")
            return
        if not _validate_md5_hex32(md5):
            QMessageBox.critical(
                self, "Preset values", "desiredConfigurationMd5 must be 32 hex characters."
            )
            return
        if not _validate_size_numeric(size):
            QMessageBox.critical(
                self, "Preset values", "desiredConfigurationSize must be numeric."
            )
            return

        if mode not in self.presets:
            self.presets[mode] = {}
        self.presets[mode][name] = {
            "desiredConfigurationTemplate": template,
            "desiredConfigurationMd5": md5.lower(),
            "desiredConfigurationSize": size,
        }
        _save_presets(self.presets)

        self._log(f"Saved preset '{name}' for {mode}.")
        self.new_preset_name.setText("")
        self._set_mode_edit(mode)
        with QSignalBlocker(self.preset_combo):
            if self.preset_combo.findText(name) >= 0:
                self.preset_combo.setCurrentText(name)
        self._per_mode_selected_preset[mode] = name
        QMessageBox.information(self, "Preset saved", f"Saved preset '{name}' for {mode}.")

    # -----------------------------
    # Sheet selector (xlsx multi-sheet)
    # -----------------------------
    def _on_sheet_changed(self, sheet_name: str):
        if self._building or not sheet_name:
            return
        self._refresh_preview_for_current_file()

    def _on_no_header_changed(self, _checked: bool):
        if self._building:
            return
        self._refresh_preview_for_current_file()

    def _selected_sheet(self) -> str | None:
        if self.sheet_combo.isVisible():
            t = self.sheet_combo.currentText()
            return t if t else None
        return None

    def _refresh_preview_for_current_file(self):
        """Re-run preview with current sheet/no_header settings, without resetting sheet combo."""
        f = self._selected_file()
        if not f:
            return
        no_header = self.no_header_cb.isChecked()
        sheet_name = self._selected_sheet()
        try:
            info = _read_input_preview(f, sheet_name=sheet_name, no_header=no_header, max_rows=5)
        except Exception as e:
            self._log(f"Preview error: {e}")
            return
        self._render_preview(f, info, update_sheet_combo=False)

    # -----------------------------
    # Preview + mapping
    # -----------------------------
    def _customer_key(self) -> str:
        return (self.customer_edit.text() or "").strip()

    def _load_customer_mapping(self) -> dict | None:
        customer = self._customer_key()
        if not customer:
            return None
        self.mappings = _load_mappings()
        v = self.mappings.get(customer)
        if isinstance(v, dict):
            return {
                "serial": str(v.get("serial", "")).strip(),
                "mac": str(v.get("mac", "")).strip(),
            }
        return None

    def _preview_selected(self):
        f = self._selected_file()
        if not f:
            self._clear_preview()
            return

        no_header = self.no_header_cb.isChecked()
        # On initial preview, use whatever sheet is currently selected (or None = active sheet)
        sheet_name = self._selected_sheet()

        try:
            info = _read_input_preview(f, sheet_name=sheet_name, no_header=no_header, max_rows=5)
        except Exception as e:
            self._log(f"Preview error for {f.name}: {e}")
            self.preview_info.setText(f"Preview error: {e}")
            return

        # Update sheet combo if xlsx with multiple sheets
        all_sheets = info.get("all_sheets", [])
        if info["type"] == "xlsx" and len(all_sheets) > 1:
            with QSignalBlocker(self.sheet_combo):
                self.sheet_combo.clear()
                self.sheet_combo.addItems(all_sheets)
                active_sheet = info["sheet"]
                if active_sheet in all_sheets:
                    self.sheet_combo.setCurrentText(active_sheet)
            self._xlsx_sheets = all_sheets
            self._set_sheet_selector_visible(True)
        else:
            self._xlsx_sheets = []
            self._set_sheet_selector_visible(False)

        self._render_preview(f, info, update_sheet_combo=False)

    def _render_preview(self, f: Path, info: dict, update_sheet_combo: bool = True):
        headers = info["headers"]

        with QSignalBlocker(self.serial_combo):
            self.serial_combo.clear()
            self.serial_combo.addItems(headers)
        with QSignalBlocker(self.mac_combo):
            self.mac_combo.clear()
            self.mac_combo.addItems(headers)

        saved = self._load_customer_mapping()
        serial_auto = _pick_column(headers, SERIAL_HEADER_ALIASES, must_contain=["serial"])
        mac_auto = _pick_column(headers, MAC_HEADER_ALIASES, must_contain=["mac"])

        mapping_src = "none"
        serial_eff = ""
        mac_eff = ""

        if saved and saved.get("serial") in headers and saved.get("mac") in headers:
            mapping_src = "saved"
            serial_eff = saved.get("serial", "")
            mac_eff = saved.get("mac", "")
        elif serial_auto and mac_auto:
            mapping_src = "auto"
            serial_eff = serial_auto
            mac_eff = mac_auto

        if serial_eff:
            with QSignalBlocker(self.serial_combo):
                self.serial_combo.setCurrentText(serial_eff)
        if mac_eff:
            with QSignalBlocker(self.mac_combo):
                self.mac_combo.setCurrentText(mac_eff)

        no_header = self.no_header_cb.isChecked()
        if info["type"] == "csv":
            self.preview_info.setText(
                f"Preview: {f.name} (csv, delimiter '{info['delimiter']}'), "
                f"headers: {len(headers)}"
            )
        else:
            header_note = " [no header — positional cols]" if no_header else ""
            self.preview_info.setText(
                f"Preview: {f.name} (xlsx, sheet '{info['sheet']}'{header_note}), "
                f"headers: {len(headers)}"
            )

        if mapping_src == "auto":
            self.mapping_info.setText(
                f"Mapping: auto (serial '{serial_eff}', mac '{mac_eff}')"
            )
        elif mapping_src == "saved":
            self.mapping_info.setText(
                f"Mapping: saved for customer '{self._customer_key()}' "
                f"(serial '{serial_eff}', mac '{mac_eff}')"
            )
        else:
            self.mapping_info.setText("Mapping: not detected (select columns manually)")

        def trunc(v: str, n: int = 60) -> str:
            v = (v or "").strip()
            return v if len(v) <= n else v[: n - 3] + "..."

        lines = ["Headers:", "  " + " | ".join(headers), "", "Rows (first 5):"]
        rows = info["rows"]
        if not rows:
            lines.append("  (no data rows)")
        else:
            for row_index, row in rows:
                parts = [trunc(str(row.get(h, ""))) for h in headers]
                lines.append(f"{row_index}: " + " | ".join(parts))

        self.preview_text.setPlainText("\n".join(lines))
        self._update_config_warning()

    def _save_mapping_for_customer(self):
        customer = self._customer_key()
        if not customer:
            QMessageBox.critical(self, "Customer mapping", "Filename input name (customer) is empty.")
            return

        f = self._selected_file()
        if not f:
            QMessageBox.critical(self, "Customer mapping", "No file selected.")
            return

        sheet_name = self._selected_sheet()
        no_header = self.no_header_cb.isChecked()
        info = _read_input_preview(f, sheet_name=sheet_name, no_header=no_header, max_rows=1)
        headers = info["headers"]

        serial_col = (self.serial_combo.currentText() or "").strip()
        mac_col = (self.mac_combo.currentText() or "").strip()

        if not serial_col or not mac_col:
            QMessageBox.critical(
                self, "Customer mapping", "Select both Serial column and MAC column."
            )
            return
        if serial_col not in headers or mac_col not in headers:
            QMessageBox.critical(
                self,
                "Customer mapping",
                "Selected columns are not in the current file headers.",
            )
            return

        self.mappings = _load_mappings()
        if customer in self.mappings:
            res = QMessageBox.question(
                self,
                "Customer mapping",
                f"Mapping for '{customer}' already exists. Overwrite?",
                QMessageBox.Yes | QMessageBox.No,
            )
            if res != QMessageBox.Yes:
                return

        self.mappings[customer] = {"serial": serial_col, "mac": mac_col}
        _save_mappings(self.mappings)

        self.mapping_info.setText(
            f"Mapping: saved for customer '{customer}' "
            f"(serial '{serial_col}', mac '{mac_col}')"
        )
        self._log(
            f"Saved mapping for customer '{customer}': serial='{serial_col}', mac='{mac_col}'"
        )
        QMessageBox.information(self, "Customer mapping", f"Saved mapping for '{customer}'.")

    # -----------------------------
    # Run scripts
    # -----------------------------
    def _validate_common_inputs(self) -> tuple[str, str]:
        if not self.files:
            raise ValueError("No input files selected.")

        date_str = (self.date_edit.text() or "").strip()
        customer = (self.customer_edit.text() or "").strip()

        if not _validate_yyyymmdd(date_str):
            raise ValueError("Filename date must be 8 digits in YYYYMMDD format.")
        if not customer:
            raise ValueError("Filename input name (customer) is empty.")
        return date_str, customer

    def _mapping_override_for_run(self) -> dict | None:
        serial_col = (self.serial_combo.currentText() or "").strip()
        mac_col = (self.mac_combo.currentText() or "").strip()
        if serial_col and mac_col:
            return {"serial": serial_col, "mac": mac_col}

        saved = self._load_customer_mapping()
        if saved and saved.get("serial") and saved.get("mac"):
            return saved
        return None

    def _run_mode(self, mode: str):
        try:
            date_str, customer = self._validate_common_inputs()

            device_type = self._per_mode_device_type.get(mode, DEVICE_VARIANTS[mode][0])

            run_cfg = _build_run_cfg(
                mode=mode,
                device_type=device_type,
                template=self.template_edit.text(),
                md5=self.md5_edit.text(),
                size=self.size_edit.text(),
            )

            # Warn if config name doesn't match device type (non-blocking)
            warning = _check_config_compatibility(
                run_cfg["desiredConfigurationTemplate"], mode, device_type
            )
            if warning:
                res = QMessageBox.question(
                    self,
                    "Config name mismatch",
                    f"{warning}\n\nContinue anyway?",
                    QMessageBox.Yes | QMessageBox.No,
                )
                if res != QMessageBox.Yes:
                    self.status.setText("Cancelled.")
                    return

            mapping_override = self._mapping_override_for_run()

            sheet_name = self._selected_sheet()
            no_header = self.no_header_cb.isChecked()

            self.status.setText(f"Running: {mode} ({device_type}) ...")
            self._log(f"Run started. Mode: {mode}, Device type: {device_type}")

            time_tag = dt.datetime.now().strftime("%H%M%S")

            for i, infile in enumerate(self.files, start=1):
                tmp_name = f"__tmp__{date_str}_{time_tag}_{i}.csv"
                tmp_path = self.output_dir / tmp_name

                data_rows = self._process_one_file(
                    infile, tmp_path, mode, run_cfg, mapping_override,
                    sheet_name=sheet_name, no_header=no_header,
                )

                final_name = _final_output_name(date_str, customer, data_rows, device_type)
                final_path = self.output_dir / final_name

                if final_path.exists():
                    tmp_path.unlink(missing_ok=True)
                    raise FileExistsError(f"Output already exists: {final_path}")

                tmp_path.replace(final_path)
                self._log(f"Saved: {final_path.name}")

            self.status.setText("Done.")
            self._log(f"Run finished successfully. Mode: {mode}, Device type: {device_type}")
        except Exception as e:
            self.status.setText("Error.")
            self._log(f"ERROR: {e}")
            QMessageBox.critical(self, f"Error [{mode}]", str(e))

    def _process_one_file(
        self,
        infile: Path,
        out_path: Path,
        mode: str,
        run_cfg: dict,
        mapping_override: dict | None,
        sheet_name: str | None = None,
        no_header: bool = False,
    ) -> int:
        self._log(f"Processing: {infile.name}")

        it = _iter_input_rows(infile, self._log, sheet_name=sheet_name, no_header=no_header)
        first = next(it, None)
        if first is None:
            raise ValueError(f"No data in file: {infile}")

        first_row_index, input_fields, first_row = first

        if mode == "Repeater":
            row_transform = _repeater_transform_factory(input_fields, run_cfg, mapping_override)
            out_fields = REPEATER_OUTPUT_FIELDS
        elif mode == "Headend":
            row_transform = _headend_transform_factory(input_fields, run_cfg, mapping_override)
            out_fields = HEADEND_OUTPUT_FIELDS
        elif mode == "Proxie":
            row_transform = _proxie_transform_factory(input_fields, run_cfg, mapping_override)
            out_fields = PROXIE_OUTPUT_FIELDS
        else:
            raise ValueError(f"Unknown mode: {mode}")

        with out_path.open("w", encoding="utf-8", newline="") as fout:
            writer = csv.DictWriter(fout, fieldnames=out_fields, delimiter=",")
            writer.writeheader()

            data_rows = 0
            skipped_empty = 0

            def handle(row_index: int, row: dict):
                nonlocal data_rows, skipped_empty
                if _is_empty_row(row):
                    skipped_empty += 1
                    self._log(f"WARNING: skipped empty row {row_index} in {infile.name}")
                    return
                out = row_transform(row, row_index)
                writer.writerow(out)
                data_rows += 1

            handle(first_row_index, first_row)
            for row_index, _fields, row in it:
                handle(row_index, row)

        self._log(
            f"OK: {infile.name} data rows written: {data_rows} (skipped empty: {skipped_empty})"
        )
        return data_rows


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet("""
        QWidget { background: #f5f5f5; color: #000000; }
        QLineEdit, QPlainTextEdit, QListWidget, QComboBox {
            background: #ffffff;
            color: #000000;
            selection-background-color: #0078d7;
            selection-color: #ffffff;
        }
        QPushButton {
            background: #f0f0f0;
            color: #000000;
            border: 1px solid #c8c8c8;
            padding: 4px 10px;
            border-radius: 4px;
        }
        QPushButton:hover { background: #e8e8e8; }
        QGroupBox { border: 1px solid #c8c8c8; margin-top: 8px; }
        QGroupBox::title { subcontrol-origin: margin; left: 8px; padding: 0 3px; }
        QCheckBox { background: transparent; }
    """)

    w = CsvToolWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
