# csv_tool.py
# GUI tool to transform device CSV/XLSX files and export comma-separated CSV.
# Features:
# - CSV input: auto-detect delimiter ',' or ';'
# - XLSX input: reads first sheet (active sheet)
# - Strict transforms for Repeater, Headend, Proxie
# - Empty rows are skipped with warnings
# - Presets loaded from ./presets/presets.json (next to this script) and saved there
# - Customer column mapping saved to ./presets/mappings.json
# - Input preview (first 5 rows) with detected header/delimiter and serial/mac mapping
# - Manual mapping UI (serial/mac column dropdown) and save per customer
# - Hover tooltips on important buttons

# -----------------------------
# Dependency bootstrap (openpyxl)
# -----------------------------
import os
import sys
import subprocess
import shutil
from pathlib import Path

REQUIRED_PACKAGES = ["openpyxl"]

def _run(cmd):
    subprocess.check_call(cmd)

def _real(p: Path) -> Path:
    return Path(os.path.realpath(str(p)))

def _venv_dir() -> Path:
    base = os.environ.get("LOCALAPPDATA")
    if base:
        return _real(Path(base)) / "csv_tool" / ".venv"
    return Path(__file__).resolve().parent / ".venv"

def _venv_python(venv_path: Path) -> Path:
    return venv_path / "Scripts" / "python.exe"

def _ensure_deps():
    try:
        import openpyxl  # noqa: F401
        return
    except ModuleNotFoundError:
        pass

    venv_path = _real(_venv_dir())
    venv_path.parent.mkdir(parents=True, exist_ok=True)

    py = _venv_python(venv_path)
    cfg = venv_path / "pyvenv.cfg"

    # Recreate broken venv
    if not py.exists() or not cfg.exists():
        if venv_path.exists():
            shutil.rmtree(venv_path, ignore_errors=True)
        _run([sys.executable, "-m", "venv", str(venv_path)])

        venv_path = _real(venv_path)
        py = _venv_python(venv_path)
        cfg = venv_path / "pyvenv.cfg"

        if not py.exists() or not cfg.exists():
            raise RuntimeError(f"Venv creation failed: missing pyvenv.cfg at {cfg}")

    _run([str(py), "-m", "pip", "install", "--upgrade", "pip"])
    _run([str(py), "-m", "pip", "install", *REQUIRED_PACKAGES])

    os.execv(str(py), [str(py)] + sys.argv)

_ensure_deps()

# -----------------------------
# Imports (after deps)
# -----------------------------
import csv
import json
import datetime as dt
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

from openpyxl import load_workbook


# -----------------------------
# Tooltip helper
# -----------------------------
class ToolTip:
    def __init__(self, widget, text: str, delay_ms: int = 450, wraplength: int = 420):
        self.widget = widget
        self.text = text
        self.delay_ms = delay_ms
        self.wraplength = wraplength
        self._after_id = None
        self._tip = None

        widget.bind("<Enter>", self._on_enter, add="+")
        widget.bind("<Leave>", self._on_leave, add="+")
        widget.bind("<ButtonPress>", self._on_leave, add="+")
        widget.bind("<FocusOut>", self._on_leave, add="+")

    def _on_enter(self, _event=None):
        self._schedule()

    def _on_leave(self, _event=None):
        self._unschedule()
        self.hide()

    def _schedule(self):
        self._unschedule()
        if self.text:
            self._after_id = self.widget.after(self.delay_ms, self.show)

    def _unschedule(self):
        if self._after_id is not None:
            try:
                self.widget.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None

    def show(self):
        if self._tip is not None or not self.text:
            return

        try:
            x = self.widget.winfo_rootx() + 12
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 10
        except Exception:
            return

        self._tip = tk.Toplevel(self.widget)
        self._tip.wm_overrideredirect(True)
        self._tip.wm_geometry(f"+{x}+{y}")

        frame = tk.Frame(self._tip, borderwidth=1, relief="solid", background="#ffffe0")
        frame.pack(fill="both", expand=True)

        label = tk.Label(
            frame,
            text=self.text,
            justify="left",
            background="#ffffe0",
            foreground="black",
            wraplength=self.wraplength,
            padx=8,
            pady=6,
        )
        label.pack()

    def hide(self):
        if self._tip is not None:
            try:
                self._tip.destroy()
            except Exception:
                pass
            self._tip = None


# -----------------------------
# Output schemas
# -----------------------------
REPEATER_OUTPUT_FIELDS = [
    "serialNumber",
    "macAddress",
    "type",
    "registrationStatus",
    "desiredConfigurationTemplate",
    "desiredConfigurationMd5",
    "desiredConfigurationSize",
]

HEADEND_OUTPUT_FIELDS = [
    "serialNumber",
    "macAddress",
    "type",
    "registrationStatus",
    "desiredConfigurationTemplate",
    "desiredConfigurationMd5",
    "desiredConfigurationSize",
]

PROXIE_OUTPUT_FIELDS = [
    "serialNumber",
    "macAddress",
    "type",
    "registrationStatus",
    "accessToken",
    "desiredConfigurationTemplate",
    "desiredConfigurationMd5",
    "desiredConfigurationSize",
]


# -----------------------------
# Fixed per-mode values
# -----------------------------
MODE_FIXED = {
    "Repeater": {"type": "R310", "registrationStatus": "ACTIVATED"},
    "Headend": {"type": "M300", "registrationStatus": "ACTIVATED"},
    "Proxie": {"type": "P300", "registrationStatus": "ACTIVATED"},
}

ACCESS_TOKEN_PREFIX = "00185803"


# -----------------------------
# Column name variants (normalized matching)
# -----------------------------
SERIAL_HEADER_ALIASES = [
    "serialnumber",
    "serialno",
    "serialnr",
    "serialnum",
    "serial",
    "sn",
    "sno",
    "deviceserialnumber",
    "devserialnumber",
    "deviceid",
    "devicenumber",
    "seriennummer",
    "seriennr",
    "sernr",
    "idserial",
    "meterserial",
]

MAC_HEADER_ALIASES = [
    "macaddress",
    "macadress",
    "macaddr",
    "mac",
    "macid",
    "maceth",
    "hwaddress",
    "hardwareaddress",
    "ethernetaddress",
    "ethmac",
    "lanmac",
    "wlanmac",
    "eui",
    "eui64",
    "mac_eth",
    "macethernet",
]


# -----------------------------
# Paths: presets and mappings stored next to script
# -----------------------------
def _script_dir() -> Path:
    return Path(__file__).resolve().parent

def _preset_folder() -> Path:
    p = _script_dir() / "presets"
    p.mkdir(parents=True, exist_ok=True)
    return p

def _presets_json_path() -> Path:
    return _preset_folder() / "presets.json"

def _mappings_json_path() -> Path:
    return _preset_folder() / "mappings.json"


# -----------------------------
# Helpers
# -----------------------------
def _norm_header(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").strip().lower())

def _pick_column(fieldnames: list[str], aliases: list[str], must_contain: list[str]) -> str | None:
    norm_map = {_norm_header(fn): fn for fn in fieldnames}

    for a in aliases:
        if a in norm_map:
            return norm_map[a]

    for fn in fieldnames:
        n = _norm_header(fn)
        if all(token in n for token in must_contain):
            return fn

    return None

def _sanitize_filename_part(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", s)
    s = s.strip().strip(".")
    return s or "input"

def _validate_yyyymmdd(s: str) -> bool:
    return bool(re.fullmatch(r"\d{8}", (s or "").strip()))

def _final_output_name(date_yyyymmdd: str, inputname: str, data_rows: int, device_type: str) -> str:
    date_part = _sanitize_filename_part(date_yyyymmdd)
    name_part = _sanitize_filename_part(inputname)
    type_part = _sanitize_filename_part(device_type)
    return f"Device_import_{date_part}_{name_part}_{data_rows}_{type_part}.csv"

def _validate_md5_hex32(s: str) -> bool:
    return bool(re.fullmatch(r"[0-9a-fA-F]{32}", (s or "").strip()))

def _validate_size_numeric(s: str) -> bool:
    return bool(re.fullmatch(r"\d+", (s or "").strip()))

def _detect_input_delimiter(file_obj) -> str:
    sample = file_obj.read(8192)
    file_obj.seek(0)

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        if getattr(dialect, "delimiter", None) in (",", ";"):
            return dialect.delimiter
    except csv.Error:
        pass

    lines = sample.splitlines()
    lines = [l for l in lines if l.strip()][:10]
    text = "\n".join(lines)
    return ";" if text.count(";") > text.count(",") else ","

def _is_empty_row(row: dict) -> bool:
    return all(((v or "").strip() == "") for v in row.values())

def _serial_c_to_b(serial: str) -> str:
    s = (serial or "").strip()
    if not s:
        raise ValueError("Headend mode: serialNumber is empty.")
    if s[0] in ("C", "c"):
        return "B" + s[1:]
    if s[0] in ("B", "b"):
        return "B" + s[1:]
    raise ValueError(f"Headend mode: serialNumber does not start with C or B: '{s}'")

def _mac_minus_one(mac: str) -> str:
    raw = re.sub(r"[^0-9a-fA-F]", "", (mac or "").strip())
    if len(raw) != 12:
        raise ValueError(f"Headend mode: invalid MAC (expected 12 hex digits): '{mac}'")
    value = int(raw, 16)
    if value == 0:
        raise ValueError(f"Headend mode: MAC underflow on decrement: '{mac}'")
    value -= 1
    h = f"{value:012X}"
    return ":".join(h[i:i + 2] for i in range(0, 12, 2))

def _normalize_mac_12hex_lower(mac: str) -> str:
    raw = re.sub(r"[^0-9a-fA-F]", "", (mac or "").strip())
    if len(raw) != 12:
        raise ValueError(f"Proxie mode: invalid MAC (expected 12 hex digits): '{mac}'")
    return raw.lower()

def _access_token_from_mac(mac: str) -> str:
    return ACCESS_TOKEN_PREFIX + _normalize_mac_12hex_lower(mac)

def _build_run_cfg(mode: str, template: str, md5: str, size: str) -> dict:
    if mode not in MODE_FIXED:
        raise ValueError(f"Unknown mode: {mode}")

    template = (template or "").strip()
    md5 = (md5 or "").strip()
    size = (size or "").strip()

    if not template:
        raise ValueError("desiredConfigurationTemplate is empty.")
    if not md5:
        raise ValueError("desiredConfigurationMd5 is empty.")
    if not size:
        raise ValueError("desiredConfigurationSize is empty.")
    if not _validate_md5_hex32(md5):
        raise ValueError("desiredConfigurationMd5 must be 32 hex characters.")
    if not _validate_size_numeric(size):
        raise ValueError("desiredConfigurationSize must be numeric.")

    return {
        "type": MODE_FIXED[mode]["type"],
        "registrationStatus": MODE_FIXED[mode]["registrationStatus"],
        "desiredConfigurationTemplate": template,
        "desiredConfigurationMd5": md5.lower(),
        "desiredConfigurationSize": size,
    }

def _read_xlsx_headers_and_rows(path: Path, max_rows: int | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if header is None:
            raise ValueError(f"No header found in file: {path}")
        fieldnames = [str(h).strip() if h is not None else "" for h in header]
        if all(h == "" for h in fieldnames):
            raise ValueError(f"Header row is empty in file: {path}")

        rows = []
        for excel_row_num, values in enumerate(it, start=2):
            row = {}
            for i, key in enumerate(fieldnames):
                v = values[i] if i < len(values) else None
                row[key] = "" if v is None else str(v).strip()
            rows.append((excel_row_num, row))
            if max_rows is not None and len(rows) >= max_rows:
                break

        return ws.title, fieldnames, rows
    finally:
        wb.close()

def _read_input_preview(infile: Path, max_rows: int = 5) -> dict:
    suffix = infile.suffix.lower()

    if suffix == ".csv":
        with infile.open("r", encoding="utf-8-sig", newline="") as fin:
            delim = _detect_input_delimiter(fin)
            reader = csv.DictReader(fin, delimiter=delim)
            if reader.fieldnames is None:
                raise ValueError(f"No header found in file: {infile}")
            headers = list(reader.fieldnames)
            rows = []
            for row_index, row in enumerate(reader, start=2):
                rows.append((row_index, row))
                if len(rows) >= max_rows:
                    break
            return {
                "type": "csv",
                "delimiter": delim,
                "sheet": "",
                "headers": headers,
                "rows": rows,
            }

    if suffix == ".xlsx":
        sheet, headers, rows = _read_xlsx_headers_and_rows(infile, max_rows=max_rows)
        return {
            "type": "xlsx",
            "delimiter": "",
            "sheet": sheet,
            "headers": headers,
            "rows": rows,
        }

    raise ValueError(f"Unsupported input type: {infile.suffix}")

def _iter_input_rows(infile: Path, log_fn):
    suffix = infile.suffix.lower()

    if suffix == ".csv":
        with infile.open("r", encoding="utf-8-sig", newline="") as fin:
            in_delim = _detect_input_delimiter(fin)
            log_fn(f"Detected input delimiter: '{in_delim}' for {infile.name}")
            reader = csv.DictReader(fin, delimiter=in_delim)
            if reader.fieldnames is None:
                raise ValueError(f"No header found in file: {infile}")
            fieldnames = list(reader.fieldnames)
            for row_index, row in enumerate(reader, start=2):
                yield row_index, fieldnames, row
        return

    if suffix == ".xlsx":
        wb = load_workbook(infile, read_only=True, data_only=True)
        try:
            ws = wb.active
            log_fn(f"Detected input type: xlsx (sheet '{ws.title}') for {infile.name}")
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if header is None:
                raise ValueError(f"No header found in file: {infile}")
            fieldnames = [str(h).strip() if h is not None else "" for h in header]
            if all(h == "" for h in fieldnames):
                raise ValueError(f"Header row is empty in file: {infile}")

            for excel_row_num, values in enumerate(it, start=2):
                row = {}
                for i, key in enumerate(fieldnames):
                    v = values[i] if i < len(values) else None
                    row[key] = "" if v is None else str(v).strip()
                yield excel_row_num, fieldnames, row
        finally:
            wb.close()
        return

    raise ValueError(f"Unsupported input type: {infile.suffix}")


# -----------------------------
# Presets: stored only in ./presets/presets.json
# -----------------------------
def _default_presets_payload() -> dict:
    return {
        "Repeater": {
            "DEFAULT": {
                "desiredConfigurationTemplate": "CFG-0001-0100-EBRO-20230223-FTPS-v1.tar.gz",
                "desiredConfigurationMd5": "f09ffd00e02bf9ad2108ea5199d46d2c",
                "desiredConfigurationSize": "195",
            }
        },
        "Headend": {
            "DEFAULT": {
                "desiredConfigurationTemplate": "CFG-0001-0100-HE-20230223-FTPS_AGC-v1.tar.gz",
                "desiredConfigurationMd5": "648be952a640f0d00f5da9f12854477f",
                "desiredConfigurationSize": "215",
            }
        },
        "Proxie": {
            "DEFAULT": {
                "desiredConfigurationTemplate": "CFG-0001-0100-PROXY-20230131-FTPS-v1.tar.gz",
                "desiredConfigurationMd5": "9768cbe43fa48cec894443e8c03ed399",
                "desiredConfigurationSize": "630",
            }
        },
    }

def _load_presets() -> dict:
    path = _presets_json_path()
    if not path.exists():
        payload = _default_presets_payload()
        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        return payload

    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    out = {"Repeater": {}, "Headend": {}, "Proxie": {}}
    for mode in out.keys():
        v = data.get(mode, {})
        if isinstance(v, dict):
            out[mode] = v
    return out

def _save_presets(presets: dict) -> None:
    path = _presets_json_path()
    payload = {
        "Repeater": presets.get("Repeater", {}),
        "Headend": presets.get("Headend", {}),
        "Proxie": presets.get("Proxie", {}),
    }
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


# -----------------------------
# Mappings: stored in ./presets/mappings.json
# { "CUSTOMER": { "serial": "SN", "mac": "mac" } }
# -----------------------------
def _load_mappings() -> dict:
    path = _mappings_json_path()
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def _save_mappings(mappings: dict) -> None:
    path = _mappings_json_path()
    with path.open("w", encoding="utf-8") as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)


# -----------------------------
# Transform factories with optional mapping override
# -----------------------------
def _resolve_serial_mac_keys(input_fieldnames: list[str], mapping_override: dict | None) -> tuple[str | None, str | None, str]:
    if mapping_override:
        s = (mapping_override.get("serial") or "").strip()
        m = (mapping_override.get("mac") or "").strip()
        if s and m and s in input_fieldnames and m in input_fieldnames:
            return s, m, "manual"

    serial_key = _pick_column(input_fieldnames, SERIAL_HEADER_ALIASES, must_contain=["serial"])
    mac_key = _pick_column(input_fieldnames, MAC_HEADER_ALIASES, must_contain=["mac"])

    if serial_key and mac_key:
        return serial_key, mac_key, "auto"

    return serial_key, mac_key, "auto"

def _repeater_transform_factory(input_fieldnames: list[str], run_cfg: dict, mapping_override: dict | None):
    serial_key, mac_key, _src = _resolve_serial_mac_keys(input_fieldnames, mapping_override)

    if not serial_key or not mac_key:
        raise ValueError(
            "Repeater mode: could not find required columns (serial and mac). "
            "Use the Mapping UI to select columns and save mapping for the customer. "
            f"Found headers: {input_fieldnames}"
        )

    def transform(row: dict, row_index: int) -> dict:
        serial = (row.get(serial_key) or "").strip()
        mac = (row.get(mac_key) or "").strip()

        if not serial:
            raise ValueError(f"Repeater mode: empty serialNumber at row {row_index} (column: '{serial_key}')")
        if not mac:
            raise ValueError(f"Repeater mode: empty macAddress at row {row_index} (column: '{mac_key}')")

        return {
            "serialNumber": serial,
            "macAddress": mac,
            "type": run_cfg["type"],
            "registrationStatus": run_cfg["registrationStatus"],
            "desiredConfigurationTemplate": run_cfg["desiredConfigurationTemplate"],
            "desiredConfigurationMd5": run_cfg["desiredConfigurationMd5"],
            "desiredConfigurationSize": run_cfg["desiredConfigurationSize"],
        }

    return transform

def _headend_transform_factory(input_fieldnames: list[str], run_cfg: dict, mapping_override: dict | None):
    serial_key, mac_key, _src = _resolve_serial_mac_keys(input_fieldnames, mapping_override)

    if not serial_key or not mac_key:
        raise ValueError(
            "Headend mode: could not find required columns (serial and mac). "
            "Use the Mapping UI to select columns and save mapping for the customer. "
            f"Found headers: {input_fieldnames}"
        )

    def transform(row: dict, row_index: int) -> dict:
        in_serial = (row.get(serial_key) or "").strip()
        in_mac = (row.get(mac_key) or "").strip()

        if not in_serial:
            raise ValueError(f"Headend mode: empty serialNumber at row {row_index} (column: '{serial_key}')")
        if not in_mac:
            raise ValueError(f"Headend mode: empty macAddress at row {row_index} (column: '{mac_key}')")

        out_serial = _serial_c_to_b(in_serial)
        out_mac = _mac_minus_one(in_mac)

        return {
            "serialNumber": out_serial,
            "macAddress": out_mac,
            "type": run_cfg["type"],
            "registrationStatus": run_cfg["registrationStatus"],
            "desiredConfigurationTemplate": run_cfg["desiredConfigurationTemplate"],
            "desiredConfigurationMd5": run_cfg["desiredConfigurationMd5"],
            "desiredConfigurationSize": run_cfg["desiredConfigurationSize"],
        }

    return transform

def _proxie_transform_factory(input_fieldnames: list[str], run_cfg: dict, mapping_override: dict | None):
    serial_key, mac_key, _src = _resolve_serial_mac_keys(input_fieldnames, mapping_override)

    if not serial_key or not mac_key:
        raise ValueError(
            "Proxie mode: could not find required columns (serial and mac). "
            "Use the Mapping UI to select columns and save mapping for the customer. "
            f"Found headers: {input_fieldnames}"
        )

    def transform(row: dict, row_index: int) -> dict:
        serial = (row.get(serial_key) or "").strip()
        mac_in = (row.get(mac_key) or "").strip()

        if not serial:
            raise ValueError(f"Proxie mode: empty serialNumber at row {row_index} (column: '{serial_key}')")
        if not mac_in:
            raise ValueError(f"Proxie mode: empty macAddress at row {row_index} (column: '{mac_key}')")

        token = _access_token_from_mac(mac_in)

        return {
            "serialNumber": serial,
            "macAddress": mac_in.strip(),
            "type": run_cfg["type"],
            "registrationStatus": run_cfg["registrationStatus"],
            "accessToken": token,
            "desiredConfigurationTemplate": run_cfg["desiredConfigurationTemplate"],
            "desiredConfigurationMd5": run_cfg["desiredConfigurationMd5"],
            "desiredConfigurationSize": run_cfg["desiredConfigurationSize"],
        }

    return transform


class CsvToolApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CSV Tool")
        self.geometry("1200x760")
        self.minsize(1200, 760)

        self.files: list[Path] = []

        self.script_dir = _script_dir()
        self.output_dir: Path = self.script_dir

        today = dt.datetime.now().strftime("%Y%m%d")
        self.date_var = tk.StringVar(value=today)
        self.inputname_var = tk.StringVar(value="")

        self.presets = _load_presets()
        self.cfg_mode_var = tk.StringVar(value="Repeater")
        self.preset_vars = {
            "Repeater": tk.StringVar(value=self._default_preset_name("Repeater")),
            "Headend": tk.StringVar(value=self._default_preset_name("Headend")),
            "Proxie": tk.StringVar(value=self._default_preset_name("Proxie")),
        }

        self.cfg_template_vars = {m: tk.StringVar(value="") for m in ("Repeater", "Headend", "Proxie")}
        self.cfg_md5_vars = {m: tk.StringVar(value="") for m in ("Repeater", "Headend", "Proxie")}
        self.cfg_size_vars = {m: tk.StringVar(value="") for m in ("Repeater", "Headend", "Proxie")}

        self._applying_preset = False
        self._apply_preset_to_vars("Repeater", self.preset_vars["Repeater"].get())
        self._apply_preset_to_vars("Headend", self.preset_vars["Headend"].get())
        self._apply_preset_to_vars("Proxie", self.preset_vars["Proxie"].get())

        self.new_preset_name_var = tk.StringVar(value="")

        self.manual_serial_col_var = tk.StringVar(value="")
        self.manual_mac_col_var = tk.StringVar(value="")
        self.preview_info_var = tk.StringVar(value="No file previewed.")
        self.mapping_info_var = tk.StringVar(value="Mapping: none")

        self._attach_custom_traces()
        self._build_ui()

    def _default_preset_name(self, mode: str) -> str:
        if mode in self.presets and isinstance(self.presets[mode], dict) and self.presets[mode]:
            if "DEFAULT" in self.presets[mode]:
                return "DEFAULT"
            return sorted(self.presets[mode].keys())[0]
        return "DEFAULT"

    def _attach_custom_traces(self):
        def make_cb(mode):
            def _cb(*_):
                if self._applying_preset:
                    return
                if self.preset_vars[mode].get() != "Custom":
                    self.preset_vars[mode].set("Custom")
                    if self.cfg_mode_var.get() == mode and hasattr(self, "preset_combo"):
                        self.preset_combo.set("Custom")
            return _cb

        for mode in ("Repeater", "Headend", "Proxie"):
            self.cfg_template_vars[mode].trace_add("write", make_cb(mode))
            self.cfg_md5_vars[mode].trace_add("write", make_cb(mode))
            self.cfg_size_vars[mode].trace_add("write", make_cb(mode))

    def _build_ui(self):
        top = tk.Frame(self)
        top.pack(fill="both", expand=False, padx=10, pady=10)

        left = tk.Frame(top)
        left.pack(side="left", fill="both", expand=True)

        tk.Label(left, text="Input files (.csv or .xlsx)").pack(anchor="w")
        self.listbox = tk.Listbox(left, height=10, selectmode=tk.EXTENDED)
        self.listbox.pack(fill="both", expand=True, pady=(4, 0))
        self.listbox.bind("<<ListboxSelect>>", self.on_file_select)

        right = tk.Frame(top)
        right.pack(side="left", fill="y", padx=(10, 0))

        btn_add = tk.Button(right, text="Add files", command=self.add_files, width=18)
        btn_add.pack(pady=(0, 6))
        ToolTip(btn_add, "Add one or more input files (.csv or .xlsx). The newly added file is auto-selected and previewed.")

        btn_remove = tk.Button(right, text="Remove selected", command=self.remove_selected, width=18)
        btn_remove.pack(pady=(0, 6))
        ToolTip(btn_remove, "Remove the selected file(s) from the list.")

        btn_clear = tk.Button(right, text="Clear list", command=self.clear_list, width=18)
        btn_clear.pack(pady=(0, 6))
        ToolTip(btn_clear, "Remove all files from the list.")

        btn_preview = tk.Button(right, text="Preview selected", command=self.preview_selected, width=18)
        btn_preview.pack(pady=(16, 6))
        ToolTip(btn_preview, "Show the first 5 rows, detected delimiter/sheet, and detected Serial/MAC columns for the selected file.")

        mid = tk.Frame(self)
        mid.pack(fill="x", expand=False, padx=10, pady=(0, 10))

        out_row = tk.Frame(mid)
        out_row.pack(fill="x", pady=(0, 8))
        tk.Label(out_row, text="Output folder").pack(side="left")

        self.out_dir_var = tk.StringVar(value=str(self.output_dir))
        tk.Entry(out_row, textvariable=self.out_dir_var, state="readonly").pack(side="left", fill="x", expand=True, padx=8)

        btn_choose = tk.Button(out_row, text="Choose", command=self.choose_output_dir, width=10)
        btn_choose.pack(side="left")
        ToolTip(btn_choose, "Select the folder where output CSV files will be written (output is always comma-separated).")

        name_row = tk.Frame(mid)
        name_row.pack(fill="x", pady=(8, 0))

        tk.Label(name_row, text="Filename date (YYYYMMDD)").pack(side="left")
        tk.Entry(name_row, textvariable=self.date_var, width=12).pack(side="left", padx=8)

        tk.Label(name_row, text="Filename input name (customer)").pack(side="left", padx=(18, 0))
        tk.Entry(name_row, textvariable=self.inputname_var, width=30).pack(side="left", padx=8)

        tk.Label(mid, text="Output filename: Device_import_YYYYMMDD_INPUTNAME_ROWCOUNT_TYPE.csv", fg="gray").pack(anchor="w", pady=(6, 0))

        cfg = tk.LabelFrame(mid, text="Desired configuration (presets + editable)")
        cfg.pack(fill="x", pady=(10, 0))

        row0 = tk.Frame(cfg)
        row0.pack(fill="x", padx=8, pady=(6, 2))
        tk.Label(row0, text="Edit for").pack(side="left")
        tk.OptionMenu(row0, self.cfg_mode_var, "Repeater", "Headend", "Proxie", command=self._on_cfg_mode_change).pack(side="left", padx=8)

        tk.Label(row0, text="Preset").pack(side="left", padx=(18, 0))
        self.preset_combo = ttk.Combobox(row0, state="readonly", width=28)
        self.preset_combo.pack(side="left", padx=8)
        self.preset_combo.bind("<<ComboboxSelected>>", self._on_preset_selected)
        ToolTip(self.preset_combo, "Select a preset to fill Template/MD5/Size. Editing any field switches to 'Custom'.")

        tk.Label(row0, text="New preset name").pack(side="left", padx=(18, 0))
        tk.Entry(row0, textvariable=self.new_preset_name_var, width=24).pack(side="left", padx=8)

        btn_save_preset = tk.Button(row0, text="Save preset", command=self.save_preset, width=12)
        btn_save_preset.pack(side="left")
        ToolTip(btn_save_preset, "Save the current Template/MD5/Size as a new named preset for the selected mode. Name must be unique.")

        row1 = tk.Frame(cfg)
        row1.pack(fill="x", padx=8, pady=2)
        tk.Label(row1, text="desiredConfigurationTemplate", width=28, anchor="w").pack(side="left")
        self.cfg_template_entry = tk.Entry(row1, width=90)
        self.cfg_template_entry.pack(side="left", fill="x", expand=True)

        row2 = tk.Frame(cfg)
        row2.pack(fill="x", padx=8, pady=2)
        tk.Label(row2, text="desiredConfigurationMd5", width=28, anchor="w").pack(side="left")
        self.cfg_md5_entry = tk.Entry(row2, width=50)
        self.cfg_md5_entry.pack(side="left", fill="x", expand=True)

        row3 = tk.Frame(cfg)
        row3.pack(fill="x", padx=8, pady=(2, 6))
        tk.Label(row3, text="desiredConfigurationSize", width=28, anchor="w").pack(side="left")
        self.cfg_size_entry = tk.Entry(row3, width=20)
        self.cfg_size_entry.pack(side="left")

        self._on_cfg_mode_change(self.cfg_mode_var.get())

        preview_frame = tk.LabelFrame(self, text="Input preview (first 5 rows) and mapping")
        preview_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        info_row = tk.Frame(preview_frame)
        info_row.pack(fill="x", padx=8, pady=(6, 4))
        tk.Label(info_row, textvariable=self.preview_info_var, fg="gray").pack(anchor="w")
        tk.Label(info_row, textvariable=self.mapping_info_var, fg="gray").pack(anchor="w", pady=(2, 0))

        mapping_row = tk.Frame(preview_frame)
        mapping_row.pack(fill="x", padx=8, pady=(6, 6))

        tk.Label(mapping_row, text="Serial column").pack(side="left")
        self.serial_combo = ttk.Combobox(mapping_row, state="readonly", width=28, textvariable=self.manual_serial_col_var)
        self.serial_combo.pack(side="left", padx=8)
        ToolTip(self.serial_combo, "Select which input column contains the device serial number (used for export).")

        tk.Label(mapping_row, text="MAC column").pack(side="left", padx=(12, 0))
        self.mac_combo = ttk.Combobox(mapping_row, state="readonly", width=28, textvariable=self.manual_mac_col_var)
        self.mac_combo.pack(side="left", padx=8)
        ToolTip(self.mac_combo, "Select which input column contains the device MAC address (used for export).")

        btn_save_mapping = tk.Button(mapping_row, text="Save mapping for customer", command=self.save_mapping, width=22)
        btn_save_mapping.pack(side="left", padx=(12, 0))
        ToolTip(btn_save_mapping, "Save the selected Serial/MAC columns for the current customer name (Filename input name). The tool reuses this mapping next time.")

        self.preview_text = tk.Text(preview_frame, height=18, wrap="none")
        self.preview_text.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        bottom = tk.Frame(self)
        bottom.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        scripts_row = tk.Frame(bottom)
        scripts_row.pack(fill="x", pady=(0, 6))

        tk.Label(scripts_row, text="Scripts").pack(side="left")

        btn_rep = tk.Button(scripts_row, text="Repeater", command=self.run_repeater, width=12)
        btn_rep.pack(side="left", padx=8)
        ToolTip(btn_rep, "Generate an output CSV for Repeaters (type R310). Uses current desired configuration for Repeater mode.")

        btn_he = tk.Button(scripts_row, text="Headend", command=self.run_headend, width=12)
        btn_he.pack(side="left", padx=8)
        ToolTip(btn_he, "Generate an output CSV for Headends (type M300). Serial: C->B. MAC: decrement by 1. Uses current desired configuration for Headend mode.")

        btn_px = tk.Button(scripts_row, text="Proxie", command=self.run_proxie, width=12)
        btn_px.pack(side="left", padx=8)
        ToolTip(btn_px, "Generate an output CSV for Proxies (type P300). Adds accessToken = 00185803 + MAC (no separators, lower-case). Uses current desired configuration for Proxie mode.")

        self.status_var = tk.StringVar(value="Ready.")
        tk.Label(scripts_row, textvariable=self.status_var, fg="gray").pack(side="left", padx=14)

        tk.Label(bottom, text="Log").pack(anchor="w")
        self.log = tk.Text(bottom, height=8, wrap="word")
        self.log.pack(fill="both", expand=True, pady=(4, 0))

        self._log_line(f"Presets file: {_presets_json_path()}")
        self._log_line(f"Mappings file: {_mappings_json_path()}")

    def _log_line(self, s: str):
        ts = dt.datetime.now().strftime("%H:%M:%S")
        self.log.insert(tk.END, f"[{ts}] {s}\n")
        self.log.see(tk.END)

    def _refresh_listbox(self):
        self.listbox.delete(0, tk.END)
        for f in self.files:
            self.listbox.insert(tk.END, str(f))

    def _select_and_preview(self, idx: int):
        if not self.files:
            return
        if idx < 0:
            idx = 0
        if idx >= len(self.files):
            idx = len(self.files) - 1

        self.listbox.selection_clear(0, tk.END)
        self.listbox.selection_set(idx)
        self.listbox.activate(idx)
        self.listbox.see(idx)

        self.preview_selected()

    def add_files(self):
        paths = filedialog.askopenfilenames(
            title="Select input files",
            filetypes=[
                ("CSV or Excel", "*.csv *.xlsx"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("All files", "*.*"),
            ],
        )
        if not paths:
            return

        added_any = False
        first_new_idx = None

        for p in map(Path, paths):
            if p not in self.files:
                self.files.append(p)
                if first_new_idx is None:
                    first_new_idx = len(self.files) - 1
                added_any = True

        self._refresh_listbox()

        if added_any:
            self._log_line(f"Added {len(paths)} file(s). Total: {len(self.files)}.")
            self._select_and_preview(first_new_idx if first_new_idx is not None else 0)
        else:
            self._log_line("No new files added (duplicates).")
            if self.files:
                self._select_and_preview(0)

    def remove_selected(self):
        sel = list(self.listbox.curselection())
        if not sel:
            return

        first_sel = sel[0]

        for idx in reversed(sel):
            del self.files[idx]

        self._refresh_listbox()
        self._log_line(f"Removed {len(sel)} file(s). Total: {len(self.files)}.")

        if self.files:
            next_idx = min(first_sel, len(self.files) - 1)
            self._select_and_preview(next_idx)
        else:
            self.preview_info_var.set("No file previewed.")
            self.mapping_info_var.set("Mapping: none")
            self.preview_text.delete("1.0", tk.END)

    def clear_list(self):
        self.files.clear()
        self._refresh_listbox()
        self._log_line("Cleared file list.")
        self.preview_info_var.set("No file previewed.")
        self.mapping_info_var.set("Mapping: none")
        self.preview_text.delete("1.0", tk.END)

    def choose_output_dir(self):
        d = filedialog.askdirectory(title="Select output folder")
        if not d:
            return
        self.output_dir = Path(d)
        self.out_dir_var.set(str(self.output_dir))
        self._log_line(f"Output folder set: {self.output_dir}")

    def on_file_select(self, _event=None):
        try:
            self.preview_selected()
        except Exception:
            pass

    def _selected_file(self) -> Path | None:
        sel = list(self.listbox.curselection())
        if not sel:
            return None
        idx = sel[0]
        if idx < 0 or idx >= len(self.files):
            return None
        return self.files[idx]

    def _preset_names_for_mode(self, mode: str) -> list[str]:
        names = []
        if mode in self.presets and isinstance(self.presets[mode], dict):
            names = sorted(self.presets[mode].keys())
        if "Custom" not in names:
            names.append("Custom")
        return names

    def _apply_preset_to_vars(self, mode: str, preset: str):
        if preset == "Custom":
            return
        values = self.presets.get(mode, {}).get(preset)
        if not isinstance(values, dict):
            return
        self._applying_preset = True
        try:
            self.cfg_template_vars[mode].set(str(values.get("desiredConfigurationTemplate", "")).strip())
            self.cfg_md5_vars[mode].set(str(values.get("desiredConfigurationMd5", "")).strip())
            self.cfg_size_vars[mode].set(str(values.get("desiredConfigurationSize", "")).strip())
        finally:
            self._applying_preset = False

    def _on_cfg_mode_change(self, mode: str):
        self.cfg_template_entry.config(textvariable=self.cfg_template_vars[mode])
        self.cfg_md5_entry.config(textvariable=self.cfg_md5_vars[mode])
        self.cfg_size_entry.config(textvariable=self.cfg_size_vars[mode])

        self.preset_combo["values"] = self._preset_names_for_mode(mode)
        self.preset_combo.config(textvariable=self.preset_vars[mode])
        self.preset_combo.set(self.preset_vars[mode].get())

    def _on_preset_selected(self, _event=None):
        mode = self.cfg_mode_var.get()
        preset = self.preset_combo.get()
        self.preset_vars[mode].set(preset)
        self._apply_preset_to_vars(mode, preset)

    def save_preset(self):
        mode = self.cfg_mode_var.get()
        name = (self.new_preset_name_var.get() or "").strip()

        if not name:
            messagebox.showerror("Preset name", "Preset name is empty.")
            return
        if name in ("DEFAULT", "Custom"):
            messagebox.showerror("Preset name", "Preset name cannot be DEFAULT or Custom.")
            return

        self.presets = _load_presets()
        if name in self.presets.get(mode, {}):
            messagebox.showerror("Preset name", f"Preset name already exists for {mode}. Choose another name.")
            return

        template = (self.cfg_template_vars[mode].get() or "").strip()
        md5 = (self.cfg_md5_vars[mode].get() or "").strip()
        size = (self.cfg_size_vars[mode].get() or "").strip()

        if not template:
            messagebox.showerror("Preset values", "desiredConfigurationTemplate is empty.")
            return
        if not _validate_md5_hex32(md5):
            messagebox.showerror("Preset values", "desiredConfigurationMd5 must be 32 hex characters.")
            return
        if not _validate_size_numeric(size):
            messagebox.showerror("Preset values", "desiredConfigurationSize must be numeric.")
            return

        preset_values = {
            "desiredConfigurationTemplate": template,
            "desiredConfigurationMd5": md5.lower(),
            "desiredConfigurationSize": size,
        }

        if mode not in self.presets:
            self.presets[mode] = {}
        self.presets[mode][name] = preset_values
        _save_presets(self.presets)

        self._on_cfg_mode_change(mode)
        self.preset_vars[mode].set(name)
        self.preset_combo.set(name)

        self._log_line(f"Saved preset '{name}' for {mode}.")
        messagebox.showinfo("Preset saved", f"Saved preset '{name}' for {mode}.")

    def _customer_key(self) -> str:
        return (self.inputname_var.get() or "").strip()

    def _load_customer_mapping(self) -> dict | None:
        customer = self._customer_key()
        if not customer:
            return None
        mappings = _load_mappings()
        v = mappings.get(customer)
        if isinstance(v, dict):
            return {"serial": str(v.get("serial", "")).strip(), "mac": str(v.get("mac", "")).strip()}
        return None

    def preview_selected(self):
        f = self._selected_file()
        if not f:
            self.preview_info_var.set("No file selected.")
            return

        info = _read_input_preview(f, max_rows=5)
        headers = info["headers"]

        self.serial_combo["values"] = headers
        self.mac_combo["values"] = headers

        saved = self._load_customer_mapping()
        serial_auto = _pick_column(headers, SERIAL_HEADER_ALIASES, must_contain=["serial"])
        mac_auto = _pick_column(headers, MAC_HEADER_ALIASES, must_contain=["mac"])

        mapping_src = "none"
        serial_eff = ""
        mac_eff = ""

        if saved and saved.get("serial") in headers and saved.get("mac") in headers:
            mapping_src = "saved"
            serial_eff = saved.get("serial", "")
            mac_eff = saved.get("mac", "")
        elif serial_auto and mac_auto:
            mapping_src = "auto"
            serial_eff = serial_auto
            mac_eff = mac_auto

        self.manual_serial_col_var.set(serial_eff)
        self.manual_mac_col_var.set(mac_eff)

        if info["type"] == "csv":
            self.preview_info_var.set(f"Preview: {f.name} (csv, delimiter '{info['delimiter']}'), headers: {len(headers)}")
        else:
            self.preview_info_var.set(f"Preview: {f.name} (xlsx, sheet '{info['sheet']}'), headers: {len(headers)}")

        if mapping_src == "auto":
            self.mapping_info_var.set(f"Mapping: auto (serial '{serial_eff}', mac '{mac_eff}')")
        elif mapping_src == "saved":
            self.mapping_info_var.set(f"Mapping: saved for customer '{self._customer_key()}' (serial '{serial_eff}', mac '{mac_eff}')")
        else:
            self.mapping_info_var.set("Mapping: not detected (select columns manually)")

        self.preview_text.delete("1.0", tk.END)
        self.preview_text.insert(tk.END, "Headers:\n")
        self.preview_text.insert(tk.END, "  " + " | ".join(headers) + "\n\n")
        self.preview_text.insert(tk.END, "Rows (first 5):\n")

        rows = info["rows"]
        if not rows:
            self.preview_text.insert(tk.END, "  (no data rows)\n")
            return

        def trunc(v: str, n: int = 50) -> str:
            v = (v or "").strip()
            return v if len(v) <= n else v[:n - 3] + "..."

        for row_index, row in rows:
            parts = []
            for h in headers:
                parts.append(trunc(str(row.get(h, ""))))
            self.preview_text.insert(tk.END, f"{row_index}: " + " | ".join(parts) + "\n")

    def save_mapping(self):
        customer = self._customer_key()
        if not customer:
            messagebox.showerror("Customer mapping", "Filename input name (customer) is empty.")
            return

        f = self._selected_file()
        if not f:
            messagebox.showerror("Customer mapping", "No file selected.")
            return

        info = _read_input_preview(f, max_rows=1)
        headers = info["headers"]

        serial_col = (self.manual_serial_col_var.get() or "").strip()
        mac_col = (self.manual_mac_col_var.get() or "").strip()

        if not serial_col or not mac_col:
            messagebox.showerror("Customer mapping", "Select both Serial column and MAC column.")
            return
        if serial_col not in headers or mac_col not in headers:
            messagebox.showerror("Customer mapping", "Selected columns are not in the current file headers.")
            return

        mappings = _load_mappings()
        if customer in mappings:
            overwrite = messagebox.askyesno("Customer mapping", f"Mapping for '{customer}' already exists. Overwrite?")
            if not overwrite:
                return

        mappings[customer] = {"serial": serial_col, "mac": mac_col}
        _save_mappings(mappings)

        self.mapping_info_var.set(f"Mapping: saved for customer '{customer}' (serial '{serial_col}', mac '{mac_col}')")
        self._log_line(f"Saved mapping for customer '{customer}': serial='{serial_col}', mac='{mac_col}'")
        messagebox.showinfo("Customer mapping", f"Saved mapping for '{customer}'.")

    def _validate_common_inputs(self):
        if not self.files:
            raise ValueError("No input files selected.")

        date_str = (self.date_var.get() or "").strip()
        inputname = (self.inputname_var.get() or "").strip()

        if not _validate_yyyymmdd(date_str):
            raise ValueError("Filename date must be 8 digits in YYYYMMDD format.")
        if not inputname:
            raise ValueError("Filename input name is empty.")
        return date_str, inputname

    def _mapping_override_for_run(self) -> dict | None:
        serial_col = (self.manual_serial_col_var.get() or "").strip()
        mac_col = (self.manual_mac_col_var.get() or "").strip()
        if serial_col and mac_col:
            return {"serial": serial_col, "mac": mac_col}

        saved = self._load_customer_mapping()
        if saved and saved.get("serial") and saved.get("mac"):
            return saved
        return None

    def run_repeater(self):
        self._run_mode("Repeater", REPEATER_OUTPUT_FIELDS, _repeater_transform_factory)

    def run_headend(self):
        self._run_mode("Headend", HEADEND_OUTPUT_FIELDS, _headend_transform_factory)

    def run_proxie(self):
        self._run_mode("Proxie", PROXIE_OUTPUT_FIELDS, _proxie_transform_factory)

    def _run_mode(self, mode: str, output_fields: list[str], transform_factory):
        try:
            date_str, inputname = self._validate_common_inputs()

            self.presets = _load_presets()

            run_cfg = _build_run_cfg(
                mode=mode,
                template=self.cfg_template_vars[mode].get(),
                md5=self.cfg_md5_vars[mode].get(),
                size=self.cfg_size_vars[mode].get(),
            )
            mapping_override = self._mapping_override_for_run()

            self.status_var.set(f"Running: {mode} ...")
            self._log_line(f"Run started (strict). Mode: {mode}")

            time_tag = dt.datetime.now().strftime("%H%M%S")
            for i, infile in enumerate(self.files, start=1):
                tmp_name = f"__tmp__{date_str}_{time_tag}_{i}.csv"
                tmp_path = self.output_dir / tmp_name

                data_rows = self._process_one_file(infile, tmp_path, output_fields, transform_factory, run_cfg, mapping_override)

                final_name = _final_output_name(date_str, inputname, data_rows, run_cfg["type"])
                final_path = self.output_dir / final_name

                if final_path.exists():
                    tmp_path.unlink(missing_ok=True)
                    raise FileExistsError(f"Output already exists: {final_path}")

                tmp_path.replace(final_path)
                self._log_line(f"Saved: {final_path.name}")

            self.status_var.set("Done.")
            self._log_line(f"Run finished successfully. Mode: {mode}")
        except Exception as e:
            self.status_var.set("Error.")
            self._log_line(f"ERROR: {e}")
            messagebox.showerror(f"Error (strict) [{mode}]", str(e))

    def _process_one_file(self, infile: Path, out_path: Path, output_fields: list[str], transform_factory, run_cfg: dict, mapping_override: dict | None):
        self._log_line(f"Processing: {infile.name}")

        it = _iter_input_rows(infile, self._log_line)
        first = next(it, None)
        if first is None:
            raise ValueError(f"No data in file: {infile}")

        first_row_index, input_fields, first_row = first
        row_transform = transform_factory(input_fields, run_cfg, mapping_override)

        with out_path.open("w", encoding="utf-8", newline="") as fout:
            writer = csv.DictWriter(fout, fieldnames=output_fields, delimiter=",")
            writer.writeheader()

            data_rows = 0
            skipped_empty = 0

            def handle(row_index: int, row: dict):
                nonlocal data_rows, skipped_empty
                if _is_empty_row(row):
                    skipped_empty += 1
                    self._log_line(f"WARNING: skipped empty row {row_index} in {infile.name}")
                    return
                out = row_transform(row, row_index)
                writer.writerow(out)
                data_rows += 1

            handle(first_row_index, first_row)
            for row_index, _fields, row in it:
                handle(row_index, row)

        self._log_line(f"OK: {infile.name} data rows written: {data_rows} (skipped empty: {skipped_empty})")
        return data_rows


if __name__ == "__main__":
    CsvToolApp().mainloop()
