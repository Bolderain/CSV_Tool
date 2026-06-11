# csv_tool_modern.py
# Modern UI/UX variant of csv_tool.py (PySide6).
#
# Same proven backend logic as csv_tool.py, but a redesigned interface:
# - Card-based single-page layout with numbered steps and a Corinex-blue accent (light mode)
# - Drag & drop files onto the list (or use the Add button)
# - One segmented mode switch (Repeater / Headend / Proxie) + one big Export button
#   (replaces the old "Edit for" dropdown + three separate run buttons)
# - Real table preview with highlighted Serial/MAC columns
# - Live OUTPUT preview tab: shows exactly what the exported rows will look like
#   (Headend C->B, MAC-1, Proxie accessToken + colon MAC) before you export
# - Inline status badges (green check when Serial/MAC auto-detected)
# - Editable device type (R310/R320/R330 ... or your own value)
# - Multi-sheet xlsx selector + "no header row" support
# - Self-bootstraps dependencies (PySide6, openpyxl) into a private venv in C:\ct\.venv
#
# The original csv_tool.py is left untouched; this is a standalone copy.

# -----------------------------
# Dependency bootstrap
# -----------------------------
import os
import sys
import subprocess
import shutil
from pathlib import Path

REQUIRED_PACKAGES = [
    "openpyxl",
    "PySide6-Essentials==6.10.2",
]


def _run(cmd: list[str]) -> None:
    subprocess.check_call(cmd)


def _real(p: Path) -> Path:
    return Path(os.path.realpath(str(p)))


def _venv_dir() -> Path:
    root = os.environ.get("SystemDrive", "C:")
    return _real(Path(root + r"\ct") / ".venv")


def _venv_python(venv_path: Path) -> Path:
    return venv_path / "Scripts" / "python.exe"


def _ensure_deps():
    try:
        import openpyxl  # noqa
        from PySide6 import QtWidgets  # noqa

        return
    except ModuleNotFoundError:
        pass

    venv_path = _real(_venv_dir())
    venv_path.parent.mkdir(parents=True, exist_ok=True)

    py = _venv_python(venv_path)
    cfg = venv_path / "pyvenv.cfg"

    if not py.exists() or not cfg.exists():
        if venv_path.exists():
            shutil.rmtree(venv_path, ignore_errors=True)
        _run([sys.executable, "-m", "venv", str(venv_path)])

        venv_path = _real(venv_path)
        py = _venv_python(venv_path)
        cfg = venv_path / "pyvenv.cfg"

        if not py.exists() or not cfg.exists():
            raise RuntimeError(f"Venv creation failed: missing pyvenv.cfg at {cfg}")

    _run([str(py), "-m", "pip", "install", "--upgrade", "pip"])
    _run([str(py), "-m", "pip", "install", *REQUIRED_PACKAGES])

    os.execv(str(py), [str(py)] + sys.argv)


_ensure_deps()

# -----------------------------
# Imports (after deps)
# -----------------------------
import csv
import json
import datetime as dt
import re

from openpyxl import load_workbook

from PySide6.QtCore import Qt, QSignalBlocker
from PySide6.QtGui import QColor, QFont
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QGroupBox,
    QVBoxLayout,
    QHBoxLayout,
    QGridLayout,
    QListWidget,
    QListWidgetItem,
    QPushButton,
    QFileDialog,
    QMessageBox,
    QLabel,
    QLineEdit,
    QComboBox,
    QPlainTextEdit,
    QScrollArea,
    QSizePolicy,
    QCheckBox,
    QTableWidget,
    QTableWidgetItem,
    QButtonGroup,
    QTabWidget,
    QHeaderView,
    QAbstractItemView,
    QFrame,
)


# -----------------------------
# Output schemas
# -----------------------------
REPEATER_OUTPUT_FIELDS = [
    "serialNumber",
    "macAddress",
    "type",
    "registrationStatus",
    "desiredConfigurationTemplate",
    "desiredConfigurationMd5",
    "desiredConfigurationSize",
]

HEADEND_OUTPUT_FIELDS = [
    "serialNumber",
    "macAddress",
    "type",
    "registrationStatus",
    "desiredConfigurationTemplate",
    "desiredConfigurationMd5",
    "desiredConfigurationSize",
]

PROXIE_OUTPUT_FIELDS = [
    "serialNumber",
    "macAddress",
    "type",
    "registrationStatus",
    "accessToken",
    "desiredConfigurationTemplate",
    "desiredConfigurationMd5",
    "desiredConfigurationSize",
]

# Device type variants per mode
DEVICE_VARIANTS = {
    "Repeater": ["R310", "R320", "R330"],
    "Headend": ["M300", "M310", "M320"],
    "Proxie": ["P300", "P310"],
}

MODES = ["Repeater", "Headend", "Proxie"]

ACCESS_TOKEN_PREFIX = "00185803"

SERIAL_HEADER_ALIASES = [
    "serialnumber",
    "serialno",
    "serialnr",
    "serialnum",
    "serial",
    "sn",
    "sno",
    "deviceserialnumber",
    "devserialnumber",
    "deviceid",
    "devicenumber",
    "seriennummer",
    "seriennr",
    "sernr",
    "idserial",
    "meterserial",
]

MAC_HEADER_ALIASES = [
    "macaddress",
    "macadress",
    "macaddr",
    "mac",
    "macid",
    "maceth",
    "hwaddress",
    "hardwareaddress",
    "ethernetaddress",
    "ethmac",
    "lanmac",
    "wlanmac",
    "eui",
    "eui64",
    "mac_eth",
    "macethernet",
]


# -----------------------------
# Paths: presets and mappings stored next to script
# -----------------------------
def _script_dir() -> Path:
    return Path(__file__).resolve().parent


def _preset_folder() -> Path:
    p = _script_dir() / "presets"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _presets_json_path() -> Path:
    return _preset_folder() / "presets.json"


def _mappings_json_path() -> Path:
    return _preset_folder() / "mappings.json"


# -----------------------------
# Helpers
# -----------------------------
def _norm_header(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").strip().lower())


def _pick_column(fieldnames: list[str], aliases: list[str], must_contain: list[str]) -> str | None:
    norm_map = {_norm_header(fn): fn for fn in fieldnames}
    for a in aliases:
        if a in norm_map:
            return norm_map[a]
    for fn in fieldnames:
        n = _norm_header(fn)
        if all(token in n for token in must_contain):
            return fn
    return None


def _sanitize_filename_part(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", s)
    s = s.strip().strip(".")
    return s or "input"


def _validate_yyyymmdd(s: str) -> bool:
    return bool(re.fullmatch(r"\d{8}", (s or "").strip()))


def _final_output_name(date_yyyymmdd: str, inputname: str, data_rows: int, device_type: str) -> str:
    date_part = _sanitize_filename_part(date_yyyymmdd)
    name_part = _sanitize_filename_part(inputname)
    type_part = _sanitize_filename_part(device_type)
    return f"Device_import_{date_part}_{name_part}_{data_rows}_{type_part}.csv"


def _validate_md5_hex32(s: str) -> bool:
    return bool(re.fullmatch(r"[0-9a-fA-F]{32}", (s or "").strip()))


def _validate_size_numeric(s: str) -> bool:
    return bool(re.fullmatch(r"\d+", (s or "").strip()))


def _detect_input_delimiter(file_obj) -> str:
    sample = file_obj.read(8192)
    file_obj.seek(0)

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        if getattr(dialect, "delimiter", None) in (",", ";"):
            return dialect.delimiter
    except csv.Error:
        pass

    lines = sample.splitlines()
    lines = [l for l in lines if l.strip()][:10]
    text = "\n".join(lines)
    return ";" if text.count(";") > text.count(",") else ","


def _is_empty_row(row: dict) -> bool:
    return all(((v or "").strip() == "") for v in row.values())


def _serial_c_to_b(serial: str) -> str:
    s = (serial or "").strip()
    if not s:
        raise ValueError("Headend mode: serialNumber is empty.")
    if s[0] in ("C", "c"):
        return "B" + s[1:]
    if s[0] in ("B", "b"):
        return "B" + s[1:]
    raise ValueError(f"Headend mode: serialNumber does not start with C or B: '{s}'")


def _mac_minus_one(mac: str) -> str:
    raw = re.sub(r"[^0-9a-fA-F]", "", (mac or "").strip())
    if len(raw) != 12:
        raise ValueError(f"Headend mode: invalid MAC (expected 12 hex digits): '{mac}'")
    value = int(raw, 16)
    if value == 0:
        raise ValueError(f"Headend mode: MAC underflow on decrement: '{mac}'")
    value -= 1
    h = f"{value:012X}"
    return ":".join(h[i : i + 2] for i in range(0, 12, 2))


def _normalize_mac_colonsep(mac: str) -> str:
    raw = re.sub(r"[^0-9a-fA-F]", "", (mac or "").strip())
    if len(raw) != 12:
        raise ValueError(f"Invalid MAC (expected 12 hex digits): '{mac}'")
    return ":".join(raw[i : i + 2].upper() for i in range(0, 12, 2))


def _normalize_mac_12hex_lower(mac: str) -> str:
    raw = re.sub(r"[^0-9a-fA-F]", "", (mac or "").strip())
    if len(raw) != 12:
        raise ValueError(f"Proxie mode: invalid MAC (expected 12 hex digits): '{mac}'")
    return raw.lower()


def _access_token_from_mac(mac: str) -> str:
    return ACCESS_TOKEN_PREFIX + _normalize_mac_12hex_lower(mac)


def _excel_col_letter(idx: int) -> str:
    """Convert 0-based column index to Excel column letter(s): 0->A, 25->Z, 26->AA ..."""
    result = ""
    n = idx + 1
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(ord("A") + r) + result
    return result


def _get_xlsx_sheet_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _check_config_compatibility(template: str, mode: str, device_type: str) -> str | None:
    """Return a warning string if template name suggests a different device type, else None."""
    if not template:
        return None
    t = template.upper()

    specific = [
        ("R310", "Repeater"), ("R320", "Repeater"), ("R330", "Repeater"),
        ("M300", "Headend"), ("M310", "Headend"), ("M320", "Headend"),
        ("P300", "Proxie"), ("P310", "Proxie"),
    ]
    for dtype, expected_mode in specific:
        if dtype in t:
            if expected_mode != mode:
                return f"Config name contains '{dtype}' ({expected_mode}) — current mode is {mode}."
            if dtype != device_type:
                return f"Config name contains '{dtype}' — selected device type is {device_type}."
            return None

    if re.search(r"(?<![A-Z])HE(?![A-Z])|HEADEND", t) and mode != "Headend":
        return f"Config name suggests Headend (HE/HEADEND) — current mode is {mode}."
    if re.search(r"PROXY|PROX(?![A-Z])", t) and mode != "Proxie":
        return f"Config name suggests Proxie (PROXY) — current mode is {mode}."
    if re.search(r"EBRO", t) and mode != "Repeater":
        return f"Config name suggests Repeater (EBRO) — current mode is {mode}."

    return None


def _build_run_cfg(mode: str, device_type: str, template: str, md5: str, size: str) -> dict:
    if mode not in DEVICE_VARIANTS:
        raise ValueError(f"Unknown mode: {mode}")

    device_type = (device_type or "").strip()
    if not device_type:
        raise ValueError(
            f"Device type is empty. Pick a preset ({', '.join(DEVICE_VARIANTS[mode])}) "
            "or type your own."
        )

    template = (template or "").strip()
    md5 = (md5 or "").strip()
    size = (size or "").strip()

    if not template:
        raise ValueError("desiredConfigurationTemplate is empty.")
    if not md5:
        raise ValueError("desiredConfigurationMd5 is empty.")
    if not size:
        raise ValueError("desiredConfigurationSize is empty.")
    if not _validate_md5_hex32(md5):
        raise ValueError("desiredConfigurationMd5 must be 32 hex characters.")
    if not _validate_size_numeric(size):
        raise ValueError("desiredConfigurationSize must be numeric.")

    return {
        "type": device_type,
        "registrationStatus": "ACTIVATED",
        "desiredConfigurationTemplate": template,
        "desiredConfigurationMd5": md5.lower(),
        "desiredConfigurationSize": size,
    }


# -----------------------------
# XLSX reading helpers
# -----------------------------
def _build_xlsx_fieldnames(raw_header: tuple, no_header: bool) -> list[str]:
    if no_header:
        return [f"Col_{_excel_col_letter(i)}" for i in range(len(raw_header))]
    result = []
    for i, h in enumerate(raw_header):
        name = str(h).strip() if h is not None else ""
        if not name:
            name = f"Col_{_excel_col_letter(i)}"
        result.append(name)
    return result


def _read_xlsx_headers_and_rows(
    path: Path,
    sheet_name: str | None = None,
    no_header: bool = False,
    max_rows: int | None = None,
):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        it = ws.iter_rows(values_only=True)
        first = next(it, None)
        if first is None:
            raise ValueError(f"No data in sheet '{ws.title}' in file: {path}")

        fieldnames = _build_xlsx_fieldnames(first, no_header)
        num_cols = len(fieldnames)

        rows = []
        if no_header:
            row0 = {
                fieldnames[i]: ("" if i >= len(first) or first[i] is None else str(first[i]).strip())
                for i in range(num_cols)
            }
            rows.append((1, row0))

        for excel_row_num, values in enumerate(it, start=2):
            row = {
                fieldnames[i]: ("" if i >= len(values) or values[i] is None else str(values[i]).strip())
                for i in range(num_cols)
            }
            rows.append((excel_row_num, row))
            if max_rows is not None and len(rows) >= max_rows:
                break

        return ws.title, fieldnames, rows
    finally:
        wb.close()


def _read_input_preview(
    infile: Path,
    sheet_name: str | None = None,
    no_header: bool = False,
    max_rows: int = 5,
) -> dict:
    suffix = infile.suffix.lower()

    if suffix == ".csv":
        with infile.open("r", encoding="utf-8-sig", newline="") as fin:
            delim = _detect_input_delimiter(fin)
            reader = csv.DictReader(fin, delimiter=delim)
            if reader.fieldnames is None:
                raise ValueError(f"No header found in file: {infile}")
            headers = list(reader.fieldnames)
            rows = []
            for row_index, row in enumerate(reader, start=2):
                rows.append((row_index, row))
                if len(rows) >= max_rows:
                    break
            return {
                "type": "csv",
                "delimiter": delim,
                "sheet": "",
                "headers": headers,
                "rows": rows,
                "all_sheets": [],
            }

    if suffix == ".xlsx":
        all_sheets = _get_xlsx_sheet_names(infile)
        sheet, headers, rows = _read_xlsx_headers_and_rows(
            infile, sheet_name=sheet_name, no_header=no_header, max_rows=max_rows
        )
        return {
            "type": "xlsx",
            "delimiter": "",
            "sheet": sheet,
            "headers": headers,
            "rows": rows,
            "all_sheets": all_sheets,
        }

    raise ValueError(f"Unsupported input type: {infile.suffix}")


def _iter_input_rows(
    infile: Path,
    log_fn,
    sheet_name: str | None = None,
    no_header: bool = False,
):
    suffix = infile.suffix.lower()

    if suffix == ".csv":
        with infile.open("r", encoding="utf-8-sig", newline="") as fin:
            in_delim = _detect_input_delimiter(fin)
            log_fn(f"Detected input delimiter: '{in_delim}' for {infile.name}")
            reader = csv.DictReader(fin, delimiter=in_delim)
            if reader.fieldnames is None:
                raise ValueError(f"No header found in file: {infile}")
            fieldnames = list(reader.fieldnames)
            for row_index, row in enumerate(reader, start=2):
                yield row_index, fieldnames, row
        return

    if suffix == ".xlsx":
        wb = load_workbook(infile, read_only=True, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            log_fn(f"Detected input type: xlsx (sheet '{ws.title}') for {infile.name}")
            it = ws.iter_rows(values_only=True)
            first = next(it, None)
            if first is None:
                raise ValueError(f"No data in sheet '{ws.title}' in file: {infile}")

            fieldnames = _build_xlsx_fieldnames(first, no_header)
            num_cols = len(fieldnames)

            if not no_header and all(fn.startswith("Col_") for fn in fieldnames):
                log_fn(
                    f"WARNING: All column headers are empty in sheet '{ws.title}'. "
                    "Consider enabling 'No header row' if the sheet has no header."
                )

            if no_header:
                row0 = {
                    fieldnames[i]: ("" if i >= len(first) or first[i] is None else str(first[i]).strip())
                    for i in range(num_cols)
                }
                yield 1, fieldnames, row0

            for excel_row_num, values in enumerate(it, start=2):
                row = {
                    fieldnames[i]: ("" if i >= len(values) or values[i] is None else str(values[i]).strip())
                    for i in range(num_cols)
                }
                yield excel_row_num, fieldnames, row
        finally:
            wb.close()
        return

    raise ValueError(f"Unsupported input type: {infile.suffix}")


# -----------------------------
# Presets / Mappings storage
# -----------------------------
def _default_presets_payload() -> dict:
    return {
        "Repeater": {
            "DEFAULT": {
                "desiredConfigurationTemplate": "CFG-0001-0100-EBRO-20230223-FTPS-v1.tar.gz",
                "desiredConfigurationMd5": "f09ffd00e02bf9ad2108ea5199d46d2c",
                "desiredConfigurationSize": "195",
            }
        },
        "Headend": {
            "DEFAULT": {
                "desiredConfigurationTemplate": "CFG-0001-0100-HE-20230223-FTPS_AGC-v1.tar.gz",
                "desiredConfigurationMd5": "648be952a640f0d00f5da9f12854477f",
                "desiredConfigurationSize": "215",
            }
        },
        "Proxie": {
            "DEFAULT": {
                "desiredConfigurationTemplate": "CFG-0001-0100-PROXY-20230131-FTPS-v1.tar.gz",
                "desiredConfigurationMd5": "9768cbe43fa48cec894443e8c03ed399",
                "desiredConfigurationSize": "630",
            }
        },
    }


def _load_presets() -> dict:
    path = _presets_json_path()
    if not path.exists():
        payload = _default_presets_payload()
        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        return payload

    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    out = {"Repeater": {}, "Headend": {}, "Proxie": {}}
    for mode in out.keys():
        v = data.get(mode, {})
        if isinstance(v, dict):
            out[mode] = v
    return out


def _save_presets(presets: dict) -> None:
    path = _presets_json_path()
    payload = {
        "Repeater": presets.get("Repeater", {}),
        "Headend": presets.get("Headend", {}),
        "Proxie": presets.get("Proxie", {}),
    }
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def _load_mappings() -> dict:
    path = _mappings_json_path()
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _save_mappings(mappings: dict) -> None:
    path = _mappings_json_path()
    with path.open("w", encoding="utf-8") as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)


# -----------------------------
# Transform factories with mapping override
# -----------------------------
def _resolve_serial_mac_keys(
    input_fieldnames: list[str], mapping_override: dict | None
) -> tuple[str | None, str | None, str]:
    if mapping_override:
        s = (mapping_override.get("serial") or "").strip()
        m = (mapping_override.get("mac") or "").strip()
        if s and m and s in input_fieldnames and m in input_fieldnames:
            return s, m, "manual"

    serial_key = _pick_column(input_fieldnames, SERIAL_HEADER_ALIASES, must_contain=["serial"])
    mac_key = _pick_column(input_fieldnames, MAC_HEADER_ALIASES, must_contain=["mac"])

    return serial_key, mac_key, "auto"


def _repeater_transform_factory(
    input_fieldnames: list[str], run_cfg: dict, mapping_override: dict | None
):
    serial_key, mac_key, _src = _resolve_serial_mac_keys(input_fieldnames, mapping_override)
    if not serial_key or not mac_key:
        raise ValueError(
            "Repeater mode: could not find required columns (serial and mac). "
            "Use the mapping dropdowns and save the mapping for the customer."
        )

    def transform(row: dict, row_index: int) -> dict:
        serial = (row.get(serial_key) or "").strip()
        mac = (row.get(mac_key) or "").strip()
        if not serial:
            raise ValueError(
                f"Repeater mode: empty serialNumber at row {row_index} (column: '{serial_key}')"
            )
        if not mac:
            raise ValueError(
                f"Repeater mode: empty macAddress at row {row_index} (column: '{mac_key}')"
            )
        return {
            "serialNumber": serial,
            "macAddress": mac,
            "type": run_cfg["type"],
            "registrationStatus": run_cfg["registrationStatus"],
            "desiredConfigurationTemplate": run_cfg["desiredConfigurationTemplate"],
            "desiredConfigurationMd5": run_cfg["desiredConfigurationMd5"],
            "desiredConfigurationSize": run_cfg["desiredConfigurationSize"],
        }

    return transform


def _headend_transform_factory(
    input_fieldnames: list[str], run_cfg: dict, mapping_override: dict | None
):
    serial_key, mac_key, _src = _resolve_serial_mac_keys(input_fieldnames, mapping_override)
    if not serial_key or not mac_key:
        raise ValueError(
            "Headend mode: could not find required columns (serial and mac). "
            "Use the mapping dropdowns and save the mapping for the customer."
        )

    def transform(row: dict, row_index: int) -> dict:
        in_serial = (row.get(serial_key) or "").strip()
        in_mac = (row.get(mac_key) or "").strip()
        if not in_serial:
            raise ValueError(
                f"Headend mode: empty serialNumber at row {row_index} (column: '{serial_key}')"
            )
        if not in_mac:
            raise ValueError(
                f"Headend mode: empty macAddress at row {row_index} (column: '{mac_key}')"
            )
        out_serial = _serial_c_to_b(in_serial)
        out_mac = _mac_minus_one(in_mac)
        return {
            "serialNumber": out_serial,
            "macAddress": out_mac,
            "type": run_cfg["type"],
            "registrationStatus": run_cfg["registrationStatus"],
            "desiredConfigurationTemplate": run_cfg["desiredConfigurationTemplate"],
            "desiredConfigurationMd5": run_cfg["desiredConfigurationMd5"],
            "desiredConfigurationSize": run_cfg["desiredConfigurationSize"],
        }

    return transform


def _proxie_transform_factory(
    input_fieldnames: list[str], run_cfg: dict, mapping_override: dict | None
):
    serial_key, mac_key, _src = _resolve_serial_mac_keys(input_fieldnames, mapping_override)
    if not serial_key or not mac_key:
        raise ValueError(
            "Proxie mode: could not find required columns (serial and mac). "
            "Use the mapping dropdowns and save the mapping for the customer."
        )

    def transform(row: dict, row_index: int) -> dict:
        serial = (row.get(serial_key) or "").strip()
        mac_in = (row.get(mac_key) or "").strip()
        if not serial:
            raise ValueError(
                f"Proxie mode: empty serialNumber at row {row_index} (column: '{serial_key}')"
            )
        if not mac_in:
            raise ValueError(
                f"Proxie mode: empty macAddress at row {row_index} (column: '{mac_key}')"
            )
        normalized_mac = _normalize_mac_colonsep(mac_in)
        token = _access_token_from_mac(mac_in)
        return {
            "serialNumber": serial,
            "macAddress": normalized_mac,
            "type": run_cfg["type"],
            "registrationStatus": run_cfg["registrationStatus"],
            "accessToken": token,
            "desiredConfigurationTemplate": run_cfg["desiredConfigurationTemplate"],
            "desiredConfigurationMd5": run_cfg["desiredConfigurationMd5"],
            "desiredConfigurationSize": run_cfg["desiredConfigurationSize"],
        }

    return transform


_FACTORY_BY_MODE = {
    "Repeater": (_repeater_transform_factory, REPEATER_OUTPUT_FIELDS),
    "Headend": (_headend_transform_factory, HEADEND_OUTPUT_FIELDS),
    "Proxie": (_proxie_transform_factory, PROXIE_OUTPUT_FIELDS),
}


# -----------------------------
# Colors (Corinex-blue light theme)
# -----------------------------
ACCENT = "#0067b8"
ACCENT_DARK = "#004f8c"
OK_GREEN = "#107c10"
WARN_ORANGE = "#b85c00"
ERR_RED = "#c42b1c"
COL_SERIAL_BG = QColor("#e3f0fb")  # light blue
COL_MAC_BG = QColor("#dff3e1")     # light green
COL_ERR_BG = QColor("#fde7e9")     # light red


# -----------------------------
# Drag & drop file list
# -----------------------------
class FileDropList(QListWidget):
    def __init__(self, on_files_dropped):
        super().__init__()
        self._on_files_dropped = on_files_dropped
        self.setAcceptDrops(True)

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()
        else:
            super().dragEnterEvent(e)

    def dragMoveEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()
        else:
            super().dragMoveEvent(e)

    def dropEvent(self, e):
        if e.mimeData().hasUrls():
            paths = [u.toLocalFile() for u in e.mimeData().urls() if u.toLocalFile()]
            if paths:
                self._on_files_dropped(paths)
            e.acceptProposedAction()
        else:
            super().dropEvent(e)


# -----------------------------
# GUI
# -----------------------------
class CsvToolModernWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Corinex Device Import Tool")
        self.resize(1290, 980)
        self.setMinimumSize(1060, 800)

        self.script_dir = _script_dir()
        self.output_dir = self.script_dir
        self.files: list[Path] = []

        self.presets = _load_presets()
        self.mappings = _load_mappings()

        self.mode = "Repeater"
        self._building = False
        self._applying_preset = False
        self._per_mode_selected_preset = {m: self._default_preset_name(m) for m in MODES}
        self._per_mode_device_type = {m: DEVICE_VARIANTS[m][0] for m in MODES}
        self._xlsx_sheets: list[str] = []
        self._sheet_selector_active = False
        self._cur_headers: list[str] = []
        self._cur_rows: list = []

        self._build_ui()
        self._log(f"Presets file: {_presets_json_path()}")
        self._log(f"Mappings file: {_mappings_json_path()}")

        self._on_mode_changed("Repeater", initial=True)
        self._update_filename_hint()

    def _default_preset_name(self, mode: str) -> str:
        block = self.presets.get(mode, {})
        if isinstance(block, dict) and block:
            if "DEFAULT" in block:
                return "DEFAULT"
            return sorted(block.keys())[0]
        return "DEFAULT"

    # -----------------------------
    # UI construction
    # -----------------------------
    def _card(self, title: str):
        box = QGroupBox(title)
        box.setObjectName("card")
        lay = QVBoxLayout(box)
        lay.setContentsMargins(14, 18, 14, 14)
        lay.setSpacing(9)
        return box, lay

    def _build_ui(self):
        self._building = True

        central = QWidget()
        root = QVBoxLayout(central)
        root.setSpacing(10)
        root.setContentsMargins(0, 0, 0, 0)

        # Header bar
        root.addWidget(self._build_header())

        # Scrollable body
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        body = QWidget()
        self.form = QVBoxLayout(body)
        self.form.setSpacing(12)
        self.form.setContentsMargins(14, 12, 14, 14)
        scroll.setWidget(body)
        root.addWidget(scroll)

        self.setCentralWidget(central)

        # Top row: two columns
        top_row = QHBoxLayout()
        top_row.setSpacing(12)
        left_col = QVBoxLayout()
        left_col.setSpacing(12)
        right_col = QVBoxLayout()
        right_col.setSpacing(12)
        top_row.addLayout(left_col, 5)
        top_row.addLayout(right_col, 6)
        self.form.addLayout(top_row)

        left_col.addWidget(self._build_files_card())
        right_col.addWidget(self._build_mode_card())
        right_col.addWidget(self._build_config_card())

        # Preview (full width)
        self.form.addWidget(self._build_preview_card())
        # Export (full width)
        self.form.addWidget(self._build_export_card())
        # Log (full width)
        self.form.addWidget(self._build_log_card())

        self._building = False

    def _build_header(self) -> QWidget:
        bar = QFrame()
        bar.setObjectName("header")
        lay = QVBoxLayout(bar)
        lay.setContentsMargins(18, 12, 18, 12)
        lay.setSpacing(2)

        title = QLabel("Corinex Device Import Tool")
        title.setObjectName("headerTitle")
        sub = QLabel("Convert customer CSV / XLSX device lists into the Corinex device import format")
        sub.setObjectName("headerSub")

        lay.addWidget(title)
        lay.addWidget(sub)
        return bar

    def _build_files_card(self) -> QWidget:
        box, lay = self._card("①  Input files")

        hint = QLabel("Drag & drop .csv / .xlsx here, or use Add files. Selecting a file previews it.")
        hint.setObjectName("muted")
        hint.setWordWrap(True)
        lay.addWidget(hint)

        self.file_list = FileDropList(self._add_paths)
        self.file_list.setToolTip("Drop files here or use Add files. Selecting a file auto-previews it.")
        self.file_list.setMinimumHeight(150)
        self.file_list.currentRowChanged.connect(self._on_file_selected)
        lay.addWidget(self.file_list)

        btn_row = QHBoxLayout()
        btn_add = QPushButton("➕  Add files")
        btn_add.setToolTip("Add one or more .csv / .xlsx files.")
        btn_add.clicked.connect(self._add_files)
        btn_remove = QPushButton("Remove")
        btn_remove.setToolTip("Remove the selected file.")
        btn_remove.clicked.connect(self._remove_selected)
        btn_clear = QPushButton("Clear")
        btn_clear.setToolTip("Remove all files.")
        btn_clear.clicked.connect(self._clear_list)
        btn_row.addWidget(btn_add)
        btn_row.addWidget(btn_remove)
        btn_row.addWidget(btn_clear)
        btn_row.addStretch(1)
        lay.addLayout(btn_row)

        return box

    def _build_mode_card(self) -> QWidget:
        box, lay = self._card("②  Device mode & type")

        # Segmented mode switch
        mode_row = QHBoxLayout()
        mode_row.setSpacing(0)
        self.mode_group = QButtonGroup(self)
        self.mode_group.setExclusive(True)
        self.mode_buttons = {}
        mode_caption = {
            "Repeater": "Repeater",
            "Headend": "Headend",
            "Proxie": "Proxie",
        }
        for m in MODES:
            b = QPushButton(mode_caption[m])
            b.setObjectName("mode")
            b.setCheckable(True)
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(lambda _checked, mm=m: self._on_mode_changed(mm))
            self.mode_group.addButton(b)
            self.mode_buttons[m] = b
            mode_row.addWidget(b)
        mode_row.addStretch(1)
        lay.addLayout(mode_row)

        self.mode_desc = QLabel("")
        self.mode_desc.setObjectName("muted")
        self.mode_desc.setWordWrap(True)
        lay.addWidget(self.mode_desc)

        # Device type
        dt_row = QHBoxLayout()
        dt_row.addWidget(QLabel("Device type"))
        self.device_type_combo = QComboBox()
        self.device_type_combo.setEditable(True)
        self.device_type_combo.setInsertPolicy(QComboBox.NoInsert)
        self.device_type_combo.setMinimumWidth(130)
        self.device_type_combo.setToolTip(
            "Pick a variant (e.g. R310 / R320 / R330) or type your own value.\n"
            "This becomes the 'type' field in the output CSV."
        )
        self.device_type_combo.currentTextChanged.connect(self._on_device_type_changed)
        dt_row.addWidget(self.device_type_combo)
        dt_row.addStretch(1)
        lay.addLayout(dt_row)

        return box

    def _build_config_card(self) -> QWidget:
        box, lay = self._card("③  Desired configuration")

        # Preset row
        preset_row = QHBoxLayout()
        preset_row.addWidget(QLabel("Preset"))
        self.preset_combo = QComboBox()
        self.preset_combo.setMinimumWidth(160)
        self.preset_combo.setToolTip("Pick a preset to fill the fields. Editing a field switches to Custom.")
        self.preset_combo.currentTextChanged.connect(self._on_preset_changed)
        preset_row.addWidget(self.preset_combo)
        preset_row.addStretch(1)
        self.new_preset_name = QLineEdit("")
        self.new_preset_name.setPlaceholderText("New preset name")
        self.new_preset_name.setToolTip("Name for a new preset (unique within this mode).")
        self.new_preset_name.setMaximumWidth(170)
        preset_row.addWidget(self.new_preset_name)
        btn_save_preset = QPushButton("Save preset")
        btn_save_preset.setToolTip("Save the current Template/MD5/Size as a new preset for this mode.")
        btn_save_preset.clicked.connect(self._save_preset)
        preset_row.addWidget(btn_save_preset)
        lay.addLayout(preset_row)

        # Fields
        grid = QGridLayout()
        grid.setColumnStretch(1, 1)
        grid.addWidget(QLabel("Template"), 0, 0)
        self.template_edit = QLineEdit("")
        self.template_edit.setToolTip("desiredConfigurationTemplate")
        self.template_edit.textChanged.connect(self._mark_custom_if_user_edit)
        grid.addWidget(self.template_edit, 0, 1, 1, 3)

        grid.addWidget(QLabel("MD5"), 1, 0)
        self.md5_edit = QLineEdit("")
        self.md5_edit.setToolTip("desiredConfigurationMd5 — 32 hex characters")
        self.md5_edit.textChanged.connect(self._mark_custom_if_user_edit)
        grid.addWidget(self.md5_edit, 1, 1, 1, 3)

        grid.addWidget(QLabel("Size"), 2, 0)
        self.size_edit = QLineEdit("")
        self.size_edit.setToolTip("desiredConfigurationSize — numeric")
        self.size_edit.setMaximumWidth(120)
        self.size_edit.textChanged.connect(self._mark_custom_if_user_edit)
        grid.addWidget(self.size_edit, 2, 1, alignment=Qt.AlignLeft)
        lay.addLayout(grid)

        self.config_warn_label = QLabel("")
        self.config_warn_label.setObjectName("warn")
        self.config_warn_label.setWordWrap(True)
        self.config_warn_label.setVisible(False)
        lay.addWidget(self.config_warn_label)

        return box

    def _build_preview_card(self) -> QWidget:
        box, lay = self._card("④  Preview & column mapping")

        # Sheet + no header
        opt_row = QHBoxLayout()
        self.sheet_label = QLabel("Sheet")
        opt_row.addWidget(self.sheet_label)
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(230)
        self.sheet_combo.setToolTip("Pick which sheet to preview & process (shown for multi-sheet Excel).")
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        opt_row.addWidget(self.sheet_combo)
        opt_row.addSpacing(16)
        self.no_header_cb = QCheckBox("No header row (first row is data)")
        self.no_header_cb.setToolTip(
            "Enable for sheets without a header row.\n"
            "Columns become Col_A, Col_B, ... — map Serial/MAC manually below."
        )
        self.no_header_cb.toggled.connect(self._on_no_header_changed)
        opt_row.addWidget(self.no_header_cb)
        opt_row.addStretch(1)
        self.sheet_label.setVisible(False)
        self.sheet_combo.setVisible(False)
        lay.addLayout(opt_row)

        # Mapping row with badges
        map_row = QHBoxLayout()
        map_row.addWidget(QLabel("Serial column"))
        self.serial_combo = QComboBox()
        self.serial_combo.setMinimumWidth(150)
        self.serial_combo.setToolTip("Which input column holds the serial number.")
        self.serial_combo.currentTextChanged.connect(self._on_mapping_combo_changed)
        map_row.addWidget(self.serial_combo)
        self.serial_badge = QLabel("")
        self.serial_badge.setMinimumWidth(74)
        map_row.addWidget(self.serial_badge)

        map_row.addSpacing(14)
        map_row.addWidget(QLabel("MAC column"))
        self.mac_combo = QComboBox()
        self.mac_combo.setMinimumWidth(150)
        self.mac_combo.setToolTip("Which input column holds the MAC address.")
        self.mac_combo.currentTextChanged.connect(self._on_mapping_combo_changed)
        map_row.addWidget(self.mac_combo)
        self.mac_badge = QLabel("")
        self.mac_badge.setMinimumWidth(74)
        map_row.addWidget(self.mac_badge)

        map_row.addSpacing(14)
        btn_save_map = QPushButton("Save mapping for customer")
        btn_save_map.setToolTip("Remember these Serial/MAC columns for the current customer name.")
        btn_save_map.clicked.connect(self._save_mapping_for_customer)
        map_row.addWidget(btn_save_map)
        map_row.addStretch(1)
        lay.addLayout(map_row)

        # Tabs: input + output preview
        self.preview_tabs = QTabWidget()
        self.input_table = self._make_table()
        self.output_table = self._make_table()
        self.preview_tabs.addTab(self.input_table, "Input (first 5 rows)")
        self.preview_tabs.addTab(self.output_table, "Output preview")
        self.preview_tabs.setMinimumHeight(230)
        lay.addWidget(self.preview_tabs)

        self.preview_info = QLabel("No file selected.")
        self.preview_info.setObjectName("muted")
        lay.addWidget(self.preview_info)

        return box

    def _make_table(self) -> QTableWidget:
        t = QTableWidget()
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.setSelectionMode(QAbstractItemView.NoSelection)
        t.verticalHeader().setVisible(False)
        t.horizontalHeader().setHighlightSections(False)
        t.horizontalHeader().setStretchLastSection(True)
        t.setAlternatingRowColors(True)
        return t

    def _build_export_card(self) -> QWidget:
        box, lay = self._card("⑤  Export")

        grid = QGridLayout()
        grid.setColumnStretch(1, 1)
        grid.addWidget(QLabel("Output folder"), 0, 0)
        self.out_dir = QLineEdit(str(self.output_dir))
        self.out_dir.setReadOnly(True)
        grid.addWidget(self.out_dir, 0, 1, 1, 3)
        btn_choose = QPushButton("Choose")
        btn_choose.setToolTip("Pick output folder. Output is always comma-separated CSV.")
        btn_choose.clicked.connect(self._choose_output_dir)
        grid.addWidget(btn_choose, 0, 4)

        grid.addWidget(QLabel("Date (YYYYMMDD)"), 1, 0)
        self.date_edit = QLineEdit(dt.datetime.now().strftime("%Y%m%d"))
        self.date_edit.setToolTip("Date used in the output filename.")
        self.date_edit.setMaximumWidth(130)
        self.date_edit.textChanged.connect(lambda _t: self._update_filename_hint())
        grid.addWidget(self.date_edit, 1, 1, alignment=Qt.AlignLeft)

        grid.addWidget(QLabel("Customer / input name"), 1, 2, alignment=Qt.AlignRight)
        self.customer_edit = QLineEdit("")
        self.customer_edit.setToolTip("Customer name — used in the filename and as the mapping key.")
        self.customer_edit.setMaximumWidth(240)
        self.customer_edit.textChanged.connect(lambda _t: self._update_filename_hint())
        grid.addWidget(self.customer_edit, 1, 3, 1, 2)
        lay.addLayout(grid)

        self.filename_hint = QLabel("")
        self.filename_hint.setObjectName("muted")
        lay.addWidget(self.filename_hint)

        run_row = QHBoxLayout()
        self.export_btn = QPushButton("▶  Export CSV")
        self.export_btn.setObjectName("primary")
        self.export_btn.setCursor(Qt.PointingHandCursor)
        self.export_btn.setToolTip("Generate the Corinex device import CSV for the selected mode.")
        self.export_btn.clicked.connect(self._export)
        run_row.addWidget(self.export_btn)
        run_row.addSpacing(16)
        self.status = QLabel("Ready.")
        self.status.setObjectName("muted")
        run_row.addWidget(self.status)
        run_row.addStretch(1)
        lay.addLayout(run_row)

        return box

    def _build_log_card(self) -> QWidget:
        box, lay = self._card("Log")
        self.log_text = QPlainTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setLineWrapMode(QPlainTextEdit.WidgetWidth)
        self.log_text.setFixedHeight(130)
        lay.addWidget(self.log_text)
        return box

    # -----------------------------
    # Logging
    # -----------------------------
    def _log(self, msg: str) -> None:
        ts = dt.datetime.now().strftime("%H:%M:%S")
        self.log_text.appendPlainText(f"[{ts}] {msg}")

    # -----------------------------
    # File list actions
    # -----------------------------
    def _add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Select input files",
            "",
            "CSV or Excel (*.csv *.xlsx);;CSV files (*.csv);;Excel files (*.xlsx);;All files (*.*)",
        )
        if paths:
            self._add_paths(paths)

    def _add_paths(self, paths: list[str]):
        first_new_index = None
        added = 0
        for p in paths:
            fp = Path(p)
            if fp.suffix.lower() not in (".csv", ".xlsx"):
                self._log(f"Skipped (unsupported type): {fp.name}")
                continue
            if fp not in self.files:
                self.files.append(fp)
                added += 1
                if first_new_index is None:
                    first_new_index = len(self.files) - 1

        self._refresh_file_list()
        self._log(f"Added {added} file(s). Total: {len(self.files)}.")
        if self.files and first_new_index is not None:
            self.file_list.setCurrentRow(first_new_index)
            self._preview_selected()

    def _remove_selected(self):
        idx = self.file_list.currentRow()
        if idx < 0 or idx >= len(self.files):
            return
        removed = self.files.pop(idx)
        self._refresh_file_list()
        self._log(f"Removed: {removed.name}. Total: {len(self.files)}.")
        if self.files:
            self.file_list.setCurrentRow(min(idx, len(self.files) - 1))
            self._preview_selected()
        else:
            self._clear_preview()

    def _clear_list(self):
        self.files.clear()
        self._refresh_file_list()
        self._log("Cleared file list.")
        self._clear_preview()

    def _refresh_file_list(self):
        with QSignalBlocker(self.file_list):
            self.file_list.clear()
            for f in self.files:
                self.file_list.addItem(QListWidgetItem(str(f)))

    def _on_file_selected(self, _row: int):
        if not self._building:
            self._preview_selected()

    def _selected_file(self) -> Path | None:
        idx = self.file_list.currentRow()
        if idx < 0 or idx >= len(self.files):
            return None
        return self.files[idx]

    def _clear_preview(self):
        self.preview_info.setText("No file selected.")
        self.input_table.clear()
        self.input_table.setRowCount(0)
        self.input_table.setColumnCount(0)
        self.output_table.clear()
        self.output_table.setRowCount(0)
        self.output_table.setColumnCount(0)
        with QSignalBlocker(self.serial_combo):
            self.serial_combo.clear()
        with QSignalBlocker(self.mac_combo):
            self.mac_combo.clear()
        self._set_badge(self.serial_badge, "none")
        self._set_badge(self.mac_badge, "none")
        self._cur_headers = []
        self._cur_rows = []
        self._xlsx_sheets = []
        self._set_sheet_selector_visible(False)

    def _set_sheet_selector_visible(self, visible: bool):
        self._sheet_selector_active = visible
        self.sheet_label.setVisible(visible)
        self.sheet_combo.setVisible(visible)

    # -----------------------------
    # Output folder
    # -----------------------------
    def _choose_output_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Select output folder", str(self.output_dir))
        if not d:
            return
        self.output_dir = Path(d)
        self.out_dir.setText(str(self.output_dir))
        self._log(f"Output folder set: {self.output_dir}")

    # -----------------------------
    # Mode switch
    # -----------------------------
    def _on_mode_changed(self, mode: str, initial: bool = False):
        self.mode = mode
        with QSignalBlocker(self.mode_buttons[mode]):
            self.mode_buttons[mode].setChecked(True)

        descs = {
            "Repeater": "Serial & MAC copied as-is. type = selected variant (R310 …).",
            "Headend": "Serial first char C→B, MAC − 1. type = selected variant (M300 …).",
            "Proxie": "MAC → colon format + accessToken = 00185803 + MAC. type = selected variant (P300 …).",
        }
        self.mode_desc.setText(descs.get(mode, ""))

        self._sync_mode_ui(mode)
        self._refresh_export_button()
        self._update_config_warning()
        self._update_filename_hint()
        if not initial:
            self._refresh_output_preview()

    def _sync_mode_ui(self, mode: str):
        """Populate preset combo + device-type combo for the given mode."""
        self.presets = _load_presets()
        names = sorted(self.presets.get(mode, {}).keys())
        if "Custom" not in names:
            names.append("Custom")
        with QSignalBlocker(self.preset_combo):
            self.preset_combo.clear()
            self.preset_combo.addItems(names)
            selected = self._per_mode_selected_preset.get(mode, "DEFAULT")
            if selected not in names:
                selected = "DEFAULT" if "DEFAULT" in names else (names[0] if names else "Custom")
            self.preset_combo.setCurrentText(selected)

        variants = DEVICE_VARIANTS.get(mode, [])
        with QSignalBlocker(self.device_type_combo):
            self.device_type_combo.clear()
            self.device_type_combo.addItems(variants)
            current_dtype = self._per_mode_device_type.get(mode, "") or (variants[0] if variants else "")
            self.device_type_combo.setCurrentText(current_dtype)
            self._per_mode_device_type[mode] = current_dtype

        self._apply_preset(mode, self.preset_combo.currentText())

    def _refresh_export_button(self):
        dtype = self._per_mode_device_type.get(self.mode, "")
        self.export_btn.setText(f"▶  Export CSV   ·   {self.mode} ({dtype})")

    # -----------------------------
    # Device type
    # -----------------------------
    def _on_device_type_changed(self, dtype: str):
        if self._building:
            return
        dtype = (dtype or "").strip()
        if not dtype:
            return
        self._per_mode_device_type[self.mode] = dtype
        self._refresh_export_button()
        self._update_config_warning()
        self._update_filename_hint()
        self._refresh_output_preview()

    # -----------------------------
    # Config warning
    # -----------------------------
    def _update_config_warning(self):
        if self._building:
            return
        template = (self.template_edit.text() or "").strip()
        dtype = self._per_mode_device_type.get(self.mode, "")
        warning = _check_config_compatibility(template, self.mode, dtype)
        if warning:
            self.config_warn_label.setText(f"⚠  {warning}")
            self.config_warn_label.setVisible(True)
        else:
            self.config_warn_label.setText("")
            self.config_warn_label.setVisible(False)

    # -----------------------------
    # Presets
    # -----------------------------
    def _apply_preset(self, mode: str, preset_name: str):
        if preset_name == "Custom":
            self._update_config_warning()
            return
        block = self.presets.get(mode, {})
        values = block.get(preset_name, {})
        if not isinstance(values, dict):
            return
        self._applying_preset = True
        try:
            with QSignalBlocker(self.template_edit):
                self.template_edit.setText(str(values.get("desiredConfigurationTemplate", "")).strip())
            with QSignalBlocker(self.md5_edit):
                self.md5_edit.setText(str(values.get("desiredConfigurationMd5", "")).strip())
            with QSignalBlocker(self.size_edit):
                self.size_edit.setText(str(values.get("desiredConfigurationSize", "")).strip())
        finally:
            self._applying_preset = False
        self._update_config_warning()
        self._refresh_output_preview()

    def _on_preset_changed(self, preset: str):
        if self._building:
            return
        self._per_mode_selected_preset[self.mode] = preset
        self._apply_preset(self.mode, preset)

    def _mark_custom_if_user_edit(self, _text: str):
        if self._building or self._applying_preset:
            return
        if self.preset_combo.currentText() != "Custom":
            with QSignalBlocker(self.preset_combo):
                if self.preset_combo.findText("Custom") >= 0:
                    self.preset_combo.setCurrentText("Custom")
            self._per_mode_selected_preset[self.mode] = "Custom"
        self._update_config_warning()
        self._refresh_output_preview()

    def _save_preset(self):
        mode = self.mode
        name = (self.new_preset_name.text() or "").strip()
        if not name:
            QMessageBox.critical(self, "Preset name", "Preset name is empty.")
            return
        if name in ("DEFAULT", "Custom"):
            QMessageBox.critical(self, "Preset name", "Preset name cannot be DEFAULT or Custom.")
            return
        self.presets = _load_presets()
        if name in self.presets.get(mode, {}):
            QMessageBox.critical(self, "Preset name", f"Preset name already exists for {mode}.")
            return

        template = (self.template_edit.text() or "").strip()
        md5 = (self.md5_edit.text() or "").strip()
        size = (self.size_edit.text() or "").strip()
        if not template:
            QMessageBox.critical(self, "Preset values", "Template is empty.")
            return
        if not _validate_md5_hex32(md5):
            QMessageBox.critical(self, "Preset values", "MD5 must be 32 hex characters.")
            return
        if not _validate_size_numeric(size):
            QMessageBox.critical(self, "Preset values", "Size must be numeric.")
            return

        if mode not in self.presets:
            self.presets[mode] = {}
        self.presets[mode][name] = {
            "desiredConfigurationTemplate": template,
            "desiredConfigurationMd5": md5.lower(),
            "desiredConfigurationSize": size,
        }
        _save_presets(self.presets)
        self._log(f"Saved preset '{name}' for {mode}.")
        self.new_preset_name.setText("")
        self._per_mode_selected_preset[mode] = name
        self._sync_mode_ui(mode)
        QMessageBox.information(self, "Preset saved", f"Saved preset '{name}' for {mode}.")

    # -----------------------------
    # Sheet / no header
    # -----------------------------
    def _on_sheet_changed(self, sheet_name: str):
        if self._building or not sheet_name:
            return
        self._refresh_preview_for_current_file()

    def _on_no_header_changed(self, _checked: bool):
        if self._building:
            return
        self._refresh_preview_for_current_file()

    def _selected_sheet(self) -> str | None:
        if self._sheet_selector_active:
            t = self.sheet_combo.currentText()
            return t if t else None
        return None

    def _refresh_preview_for_current_file(self):
        f = self._selected_file()
        if not f:
            return
        no_header = self.no_header_cb.isChecked()
        sheet_name = self._selected_sheet()
        try:
            info = _read_input_preview(f, sheet_name=sheet_name, no_header=no_header, max_rows=5)
        except Exception as e:
            self._log(f"Preview error: {e}")
            return
        self._render_preview(f, info)

    # -----------------------------
    # Customer mapping
    # -----------------------------
    def _customer_key(self) -> str:
        return (self.customer_edit.text() or "").strip()

    def _load_customer_mapping(self) -> dict | None:
        customer = self._customer_key()
        if not customer:
            return None
        self.mappings = _load_mappings()
        v = self.mappings.get(customer)
        if isinstance(v, dict):
            return {"serial": str(v.get("serial", "")).strip(), "mac": str(v.get("mac", "")).strip()}
        return None

    # -----------------------------
    # Preview
    # -----------------------------
    def _preview_selected(self):
        f = self._selected_file()
        if not f:
            self._clear_preview()
            return
        no_header = self.no_header_cb.isChecked()
        sheet_name = self._selected_sheet()
        try:
            info = _read_input_preview(f, sheet_name=sheet_name, no_header=no_header, max_rows=5)
        except Exception as e:
            self._log(f"Preview error for {f.name}: {e}")
            self.preview_info.setText(f"Preview error: {e}")
            return

        all_sheets = info.get("all_sheets", [])
        if info["type"] == "xlsx" and len(all_sheets) > 1:
            with QSignalBlocker(self.sheet_combo):
                self.sheet_combo.clear()
                self.sheet_combo.addItems(all_sheets)
                if info["sheet"] in all_sheets:
                    self.sheet_combo.setCurrentText(info["sheet"])
            self._xlsx_sheets = all_sheets
            self._set_sheet_selector_visible(True)
        else:
            self._xlsx_sheets = []
            self._set_sheet_selector_visible(False)

        self._render_preview(f, info)

    def _render_preview(self, f: Path, info: dict):
        headers = info["headers"]
        rows = info["rows"]
        self._cur_headers = headers
        self._cur_rows = rows

        # mapping combos
        with QSignalBlocker(self.serial_combo):
            self.serial_combo.clear()
            self.serial_combo.addItems(headers)
        with QSignalBlocker(self.mac_combo):
            self.mac_combo.clear()
            self.mac_combo.addItems(headers)

        saved = self._load_customer_mapping()
        serial_auto = _pick_column(headers, SERIAL_HEADER_ALIASES, must_contain=["serial"])
        mac_auto = _pick_column(headers, MAC_HEADER_ALIASES, must_contain=["mac"])

        serial_src = "none"
        mac_src = "none"
        serial_eff = ""
        mac_eff = ""

        if saved and saved.get("serial") in headers and saved.get("mac") in headers:
            serial_src = mac_src = "saved"
            serial_eff = saved.get("serial", "")
            mac_eff = saved.get("mac", "")
        else:
            if serial_auto:
                serial_src = "auto"
                serial_eff = serial_auto
            if mac_auto:
                mac_src = "auto"
                mac_eff = mac_auto

        if serial_eff:
            with QSignalBlocker(self.serial_combo):
                self.serial_combo.setCurrentText(serial_eff)
        if mac_eff:
            with QSignalBlocker(self.mac_combo):
                self.mac_combo.setCurrentText(mac_eff)

        self._set_badge(self.serial_badge, serial_src)
        self._set_badge(self.mac_badge, mac_src)

        no_header = self.no_header_cb.isChecked()
        if info["type"] == "csv":
            self.preview_info.setText(
                f"{f.name}  ·  csv (delimiter '{info['delimiter']}')  ·  {len(headers)} columns  ·  showing {len(rows)} rows"
            )
        else:
            note = "  ·  no header (positional)" if no_header else ""
            self.preview_info.setText(
                f"{f.name}  ·  xlsx sheet '{info['sheet']}'{note}  ·  {len(headers)} columns  ·  showing {len(rows)} rows"
            )

        self._fill_input_table(headers, rows)
        self._refresh_output_preview()

    def _current_serial_mac(self) -> tuple[str, str]:
        return (self.serial_combo.currentText() or "").strip(), (self.mac_combo.currentText() or "").strip()

    def _fill_input_table(self, headers: list[str], rows: list):
        serial_col, mac_col = self._current_serial_mac()
        t = self.input_table
        t.clear()
        t.setColumnCount(len(headers))
        t.setRowCount(len(rows))
        labels = []
        for h in headers:
            tag = ""
            if h == serial_col:
                tag = "  (Serial)"
            elif h == mac_col:
                tag = "  (MAC)"
            labels.append(f"{h}{tag}")
        t.setHorizontalHeaderLabels(labels)

        for r_idx, (_row_index, row) in enumerate(rows):
            for c_idx, h in enumerate(headers):
                item = QTableWidgetItem(str(row.get(h, "")))
                if h == serial_col:
                    item.setBackground(COL_SERIAL_BG)
                elif h == mac_col:
                    item.setBackground(COL_MAC_BG)
                t.setItem(r_idx, c_idx, item)
        t.resizeColumnsToContents()

    def _refresh_output_preview(self):
        if self._building:
            return
        headers = self._cur_headers
        rows = self._cur_rows
        t = self.output_table
        if not headers or not rows:
            t.clear()
            t.setRowCount(0)
            t.setColumnCount(0)
            return

        mapping_override = self._mapping_override_for_run()
        factory, out_fields = _FACTORY_BY_MODE[self.mode]

        dtype = self._per_mode_device_type.get(self.mode, "")
        md5 = (self.md5_edit.text() or "").strip()
        cfg = {
            "type": dtype or "(type)",
            "registrationStatus": "ACTIVATED",
            "desiredConfigurationTemplate": (self.template_edit.text() or "").strip() or "(template)",
            "desiredConfigurationMd5": (md5.lower() if md5 else "(md5)"),
            "desiredConfigurationSize": (self.size_edit.text() or "").strip() or "(size)",
        }

        try:
            transform = factory(headers, cfg, mapping_override)
        except ValueError as e:
            t.clear()
            t.setColumnCount(1)
            t.setRowCount(1)
            t.setHorizontalHeaderLabels(["Output preview"])
            item = QTableWidgetItem(f"⚠  {e}")
            item.setForeground(QColor(WARN_ORANGE))
            t.setItem(0, 0, item)
            t.resizeColumnsToContents()
            return

        out_rows = []
        for row_index, row in rows:
            if _is_empty_row(row):
                continue
            try:
                out_rows.append(("ok", transform(row, row_index)))
            except ValueError as e:
                out_rows.append(("err", str(e)))

        t.clear()
        t.setColumnCount(len(out_fields))
        t.setRowCount(len(out_rows))
        t.setHorizontalHeaderLabels(out_fields)
        for r_idx, (kind, payload) in enumerate(out_rows):
            if kind == "ok":
                for c_idx, field in enumerate(out_fields):
                    t.setItem(r_idx, c_idx, QTableWidgetItem(str(payload.get(field, ""))))
            else:
                item = QTableWidgetItem(f"⚠  {payload}")
                item.setForeground(QColor(ERR_RED))
                item.setBackground(COL_ERR_BG)
                t.setItem(r_idx, 0, item)
                for c_idx in range(1, len(out_fields)):
                    blank = QTableWidgetItem("")
                    blank.setBackground(COL_ERR_BG)
                    t.setItem(r_idx, c_idx, blank)
        t.resizeColumnsToContents()

    def _set_badge(self, label: QLabel, src: str):
        if src == "auto":
            label.setText("✓ auto")
            label.setStyleSheet(f"color: {OK_GREEN}; font-weight: bold;")
            label.setToolTip("Column auto-detected from its header name.")
        elif src == "saved":
            label.setText("✓ saved")
            label.setStyleSheet(f"color: {OK_GREEN}; font-weight: bold;")
            label.setToolTip("Loaded from the saved mapping for this customer.")
        elif src == "manual":
            label.setText("● manual")
            label.setStyleSheet(f"color: {ACCENT}; font-weight: bold;")
            label.setToolTip("Manually selected.")
        else:
            label.setText("⚠ pick")
            label.setStyleSheet(f"color: {WARN_ORANGE}; font-weight: bold;")
            label.setToolTip("Not detected — please pick the column manually.")

    def _on_mapping_combo_changed(self, _text: str):
        if self._building:
            return
        # Reflect manual choice in badges + tables
        serial_col, mac_col = self._current_serial_mac()
        serial_auto = _pick_column(self._cur_headers, SERIAL_HEADER_ALIASES, must_contain=["serial"])
        mac_auto = _pick_column(self._cur_headers, MAC_HEADER_ALIASES, must_contain=["mac"])
        self._set_badge(self.serial_badge, "auto" if serial_col and serial_col == serial_auto else ("manual" if serial_col else "none"))
        self._set_badge(self.mac_badge, "auto" if mac_col and mac_col == mac_auto else ("manual" if mac_col else "none"))
        if self._cur_headers:
            self._fill_input_table(self._cur_headers, self._cur_rows)
        self._refresh_output_preview()

    def _save_mapping_for_customer(self):
        customer = self._customer_key()
        if not customer:
            QMessageBox.critical(self, "Customer mapping", "Customer / input name is empty.")
            return
        f = self._selected_file()
        if not f:
            QMessageBox.critical(self, "Customer mapping", "No file selected.")
            return
        headers = self._cur_headers
        serial_col, mac_col = self._current_serial_mac()
        if not serial_col or not mac_col:
            QMessageBox.critical(self, "Customer mapping", "Select both Serial column and MAC column.")
            return
        if serial_col not in headers or mac_col not in headers:
            QMessageBox.critical(self, "Customer mapping", "Selected columns are not in the current headers.")
            return

        self.mappings = _load_mappings()
        if customer in self.mappings:
            res = QMessageBox.question(
                self, "Customer mapping",
                f"Mapping for '{customer}' already exists. Overwrite?",
                QMessageBox.Yes | QMessageBox.No,
            )
            if res != QMessageBox.Yes:
                return
        self.mappings[customer] = {"serial": serial_col, "mac": mac_col}
        _save_mappings(self.mappings)
        self._set_badge(self.serial_badge, "saved")
        self._set_badge(self.mac_badge, "saved")
        self._log(f"Saved mapping for '{customer}': serial='{serial_col}', mac='{mac_col}'")
        QMessageBox.information(self, "Customer mapping", f"Saved mapping for '{customer}'.")

    # -----------------------------
    # Filename hint
    # -----------------------------
    def _update_filename_hint(self):
        date_str = (self.date_edit.text() or "").strip() or "YYYYMMDD"
        customer = (self.customer_edit.text() or "").strip() or "CUSTOMER"
        dtype = self._per_mode_device_type.get(self.mode, "TYPE") or "TYPE"
        name = _final_output_name(date_str, customer, 0, dtype).replace("_0_", "_<rows>_")
        self.filename_hint.setText(f"Output filename:  {name}")

    # -----------------------------
    # Export
    # -----------------------------
    def _mapping_override_for_run(self) -> dict | None:
        serial_col, mac_col = self._current_serial_mac()
        if serial_col and mac_col:
            return {"serial": serial_col, "mac": mac_col}
        saved = self._load_customer_mapping()
        if saved and saved.get("serial") and saved.get("mac"):
            return saved
        return None

    def _validate_common_inputs(self) -> tuple[str, str]:
        if not self.files:
            raise ValueError("No input files selected.")
        date_str = (self.date_edit.text() or "").strip()
        customer = (self.customer_edit.text() or "").strip()
        if not _validate_yyyymmdd(date_str):
            raise ValueError("Date must be 8 digits (YYYYMMDD).")
        if not customer:
            raise ValueError("Customer / input name is empty.")
        return date_str, customer

    def _export(self):
        mode = self.mode
        try:
            date_str, customer = self._validate_common_inputs()
            device_type = self._per_mode_device_type.get(mode, DEVICE_VARIANTS[mode][0])
            run_cfg = _build_run_cfg(
                mode=mode,
                device_type=device_type,
                template=self.template_edit.text(),
                md5=self.md5_edit.text(),
                size=self.size_edit.text(),
            )

            warning = _check_config_compatibility(run_cfg["desiredConfigurationTemplate"], mode, device_type)
            if warning:
                res = QMessageBox.question(
                    self, "Config name mismatch",
                    f"{warning}\n\nContinue anyway?",
                    QMessageBox.Yes | QMessageBox.No,
                )
                if res != QMessageBox.Yes:
                    self.status.setText("Cancelled.")
                    return

            mapping_override = self._mapping_override_for_run()
            sheet_name = self._selected_sheet()
            no_header = self.no_header_cb.isChecked()

            self.status.setText(f"Running: {mode} ({device_type}) …")
            self._log(f"Export started. Mode: {mode}, Device type: {device_type}")
            time_tag = dt.datetime.now().strftime("%H%M%S")

            for i, infile in enumerate(self.files, start=1):
                tmp_name = f"__tmp__{date_str}_{time_tag}_{i}.csv"
                tmp_path = self.output_dir / tmp_name
                data_rows = self._process_one_file(
                    infile, tmp_path, mode, run_cfg, mapping_override,
                    sheet_name=sheet_name, no_header=no_header,
                )
                final_name = _final_output_name(date_str, customer, data_rows, device_type)
                final_path = self.output_dir / final_name
                if final_path.exists():
                    tmp_path.unlink(missing_ok=True)
                    raise FileExistsError(f"Output already exists: {final_path}")
                tmp_path.replace(final_path)
                self._log(f"Saved: {final_path.name}")

            self.status.setText("Done. ✓")
            self.status.setStyleSheet(f"color: {OK_GREEN}; font-weight: bold;")
            self._log(f"Export finished successfully. Mode: {mode}, Device type: {device_type}")
        except Exception as e:
            self.status.setText("Error.")
            self.status.setStyleSheet(f"color: {ERR_RED}; font-weight: bold;")
            self._log(f"ERROR: {e}")
            QMessageBox.critical(self, f"Error [{mode}]", str(e))

    def _process_one_file(
        self, infile: Path, out_path: Path, mode: str, run_cfg: dict,
        mapping_override: dict | None, sheet_name: str | None = None, no_header: bool = False,
    ) -> int:
        self._log(f"Processing: {infile.name}")
        it = _iter_input_rows(infile, self._log, sheet_name=sheet_name, no_header=no_header)
        first = next(it, None)
        if first is None:
            raise ValueError(f"No data in file: {infile}")
        first_row_index, input_fields, first_row = first

        factory, out_fields = _FACTORY_BY_MODE[mode]
        row_transform = factory(input_fields, run_cfg, mapping_override)

        with out_path.open("w", encoding="utf-8", newline="") as fout:
            writer = csv.DictWriter(fout, fieldnames=out_fields, delimiter=",")
            writer.writeheader()
            data_rows = 0
            skipped_empty = 0

            def handle(row_index: int, row: dict):
                nonlocal data_rows, skipped_empty
                if _is_empty_row(row):
                    skipped_empty += 1
                    self._log(f"WARNING: skipped empty row {row_index} in {infile.name}")
                    return
                writer.writerow(row_transform(row, row_index))
                data_rows += 1

            handle(first_row_index, first_row)
            for row_index, _fields, row in it:
                handle(row_index, row)

        self._log(f"OK: {infile.name} rows written: {data_rows} (skipped empty: {skipped_empty})")
        return data_rows


# -----------------------------
# Stylesheet
# -----------------------------
def _stylesheet() -> str:
    return f"""
    QWidget {{
        background: #eef1f5;
        color: #1b1b1b;
        font-size: 10pt;
    }}
    QFrame#header {{
        background: {ACCENT};
        border: none;
    }}
    QLabel#headerTitle {{
        color: #ffffff;
        font-size: 17pt;
        font-weight: bold;
        background: transparent;
    }}
    QLabel#headerSub {{
        color: #d7e7f6;
        font-size: 10pt;
        background: transparent;
    }}
    QLabel#muted {{ color: #5b6068; background: transparent; }}
    QLabel#warn {{ color: {WARN_ORANGE}; font-weight: bold; background: transparent; }}
    QGroupBox#card {{
        background: #ffffff;
        border: 1px solid #d6dce3;
        border-radius: 9px;
        margin-top: 12px;
        font-weight: bold;
    }}
    QGroupBox#card::title {{
        subcontrol-origin: margin;
        left: 14px; top: 1px;
        padding: 1px 6px;
        color: {ACCENT};
        font-size: 11pt;
    }}
    QLineEdit, QPlainTextEdit, QComboBox, QListWidget, QTableWidget {{
        background: #ffffff;
        color: #1b1b1b;
        border: 1px solid #c4cbd3;
        border-radius: 5px;
        padding: 4px 6px;
        selection-background-color: {ACCENT};
        selection-color: #ffffff;
    }}
    QLineEdit:focus, QComboBox:focus, QPlainTextEdit:focus {{
        border: 1px solid {ACCENT};
    }}
    QComboBox::drop-down {{ border: none; width: 18px; }}
    QPushButton {{
        background: #ffffff;
        color: #1b1b1b;
        border: 1px solid #c4cbd3;
        border-radius: 5px;
        padding: 6px 12px;
    }}
    QPushButton:hover {{ background: #f1f6fb; border-color: {ACCENT}; }}
    QPushButton:pressed {{ background: #e4eef8; }}
    QPushButton#primary {{
        background: {ACCENT};
        color: #ffffff;
        border: none;
        border-radius: 7px;
        padding: 11px 22px;
        font-size: 11pt;
        font-weight: bold;
    }}
    QPushButton#primary:hover {{ background: {ACCENT_DARK}; }}
    QPushButton#mode {{
        background: #ffffff;
        color: #1b1b1b;
        border: 1px solid #c4cbd3;
        padding: 8px 20px;
        font-weight: bold;
        border-radius: 0px;
    }}
    QPushButton#mode:hover {{ background: #f1f6fb; }}
    QPushButton#mode:checked {{
        background: {ACCENT};
        color: #ffffff;
        border-color: {ACCENT};
    }}
    QTabWidget::pane {{
        border: 1px solid #d6dce3;
        border-radius: 6px;
        top: -1px;
    }}
    QTabBar::tab {{
        background: #e6eaf0;
        color: #3a3f46;
        border: 1px solid #d6dce3;
        border-bottom: none;
        border-top-left-radius: 6px;
        border-top-right-radius: 6px;
        padding: 6px 16px;
        margin-right: 2px;
    }}
    QTabBar::tab:selected {{
        background: #ffffff;
        color: {ACCENT};
        font-weight: bold;
    }}
    QHeaderView::section {{
        background: #f2f5f9;
        color: #2a2f36;
        border: none;
        border-right: 1px solid #e1e6ec;
        border-bottom: 1px solid #d6dce3;
        padding: 5px 8px;
        font-weight: bold;
    }}
    QTableWidget {{
        gridline-color: #e7ebf0;
        alternate-background-color: #f7f9fc;
    }}
    QListWidget {{ padding: 2px; }}
    QScrollArea {{ background: #eef1f5; border: none; }}
    """


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet(_stylesheet())

    f = QFont("Segoe UI", 10)
    app.setFont(f)

    w = CsvToolModernWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
